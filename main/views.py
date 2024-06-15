from django.template.loader import render_to_string
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.core.paginator import Paginator, EmptyPage
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.contrib.auth.hashers import check_password
from django.db.models import Q, F, Value as V, Count, Sum, Case, When, IntegerField, ExpressionList
from django.db.models import TextField
from django.db.models.functions import Cast
from django.shortcuts import get_object_or_404, render
from django.urls import reverse

from rest_framework.parsers import JSONParser
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from rest_framework.authtoken.models import Token
from rest_framework.response import Response
from rest_framework import status, generics, filters
from rest_framework.exceptions import ValidationError
import traceback
import os
import re
import traceback
import logging
import io
import json
import uuid
import math
import base64
import requests
import time
import math
import datetime
from random import randint
import hashlib

import boto3
import pandas as pd
from bs4 import BeautifulSoup
from keycloak import keycloak_admin
from keycloak import keycloak_openid
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.protection import Protection
from openpyxl.styles import Font
from tools.decorators import data_response
from deviare import settings as deviare_settings
from django.template.loader import get_template

from weasyprint import HTML
import pygal
from pygal.style import BlueStyle
from .sample import user_progress_data, digital_readiness_data

from .models import (
    UserSettings,
    Company,
    Project,
    Assessment,
    Apprenticeship,
    Experiment,
    Deployment,
    Course,
    GCIndexAssessment,
    CourseLicenseUser,
    CourseUserAssignment,
    UserReport,
    TMForumUserResponse,
)
from .serializers import (
    CompanySerializer,
    UserSettingsSerializer,
    CourseSerializer,
    ProjectSerializer,
    GCIndexAssessmentSerializer,
    CourseLicenseSerializer,
    ProjectDashboardSerializer,
    CourseLicenseUserSerializer,
    CourseUserAssignmentSerializer,
    TMForumUserResponseDocSerializer,
    UserProgressReportSerializer,
    CustomerDigitalReadinessReportSerializer,
    UserLearningPathReportSerializer,
)
from .utils import email_check, randomString, return_image_extenstion,to_secs
from main.authhelper import get_token_from_code, get_signin_url, get_access_token
from main.outlookservice import get_me, get_my_messages, get_attachment
from main.countries import countries

from .tasks import (
    create_course_on_talent_lms,
    assign_user_to_course_on_talent_lms,
    add_user_to_talent_lms,
    edit_user_profile_on_talent_lms, change_user_status_to_inactive, change_user_status_to_active,
)

logger = logging.getLogger(__name__)


API_KEY = "U8ZnmuEC34FhHIn4nfHLQiJNSZTtp2"
URL_ADDRESS = "https://deviare.talentlms.com/api/v1"


def email_queue(recipient, subject, body, msg_html=None, **kwargs):
    try:
        from notification.models import EmailMessage, EmailTemplate
        from tools.email import ren_string_tmpl
        from time import time
        TO = recipient if type(recipient) is list else [recipient]
        msg = body or msg_html
        e = EmailMessage.objects.create(
            template=EmailTemplate.objects.filter(name='message').first(),
            content=dict(message=msg),
            recipient=TO[0],
            subject=subject,
            reference=f'{TO[0]}-{time()}', proc_id=2000000)
    except Exception as e:
        logger.exception(e)
        pass


def send_email(recipient, subject, body, msg_html=None):
    try:
        FROM = "noreply@deviare.co.za"
        # email_queue(recipient, subject, body, msg_html)
        TO = recipient if type(recipient) is list else [recipient]
        msg = send_mail(subject, body, FROM, TO, html_message=msg_html, fail_silently=False)
    except Exception as e:
        logger.exception(e)
        pass


# to peform keycloak auth and provide token
def keycloak_auth(username, password):

    keyobj = keycloak_openid.KeycloakOpenID(
        server_url=deviare_settings.KEYCLOAK_URL,
        client_id=deviare_settings.KEYCLOAK_CLIENTID,
        realm_name=deviare_settings.KEYCLOAK_REALM,
        client_secret_key=deviare_settings.KEYCLOAK_CLIENTSECRET,
        verify=True,
    )
    try:
        token = keyobj.token(username, password)
        token_val = token["access_token"]
        return token_val
    except Exception as exc:
        logger.exception(exc)
        return "error"


# To get token used in SSO
def get_impersonate_token(username):
    try:
        keyadmin = keycloak_admin.KeycloakAdmin(
            server_url=deviare_settings.KEYCLOAK_URL,
            username=deviare_settings.KEYCLOAK_ADMINUSER,
            password=deviare_settings.KEYCLOAK_ADMINPASSWORD,
            realm_name=deviare_settings.KEYCLOAK_REALM,
        )
        user_id_keycloak = keyadmin.get_user_id(username)
        tokens = keyadmin.impersonate(user_id_keycloak)
        return tokens
    except Exception as e:
        traceback.print_exc()
        raise ValueError("Invalid User")


# Create or get token
def create_or_get_token(user):
    token = Token.objects.get_or_create(user=user)
    for i in token:
        return i.key


def get_idptoken(username="", **kw):
    import json
    from urllib.parse import quote

    try:
        ipt = get_impersonate_token(username)
        ipt["SetCookie"] = {
            k.replace("HttpOnly, ", "HTTP__"): v
            for k, v in map(lambda s: s.split("=", 1), ipt["Set-Cookie"].split("; "))
        }
        ipt["url_search_string"] = "{HTTP__KEYCLOAK_IDENTITY}&session={HTTP__KEYCLOAK_SESSION}&url=".format(
            **{k: quote(v) for k, v in ipt["SetCookie"].items() if k.startswith("HTTP")}
        )
        jipt = json.dumps({k: str(v) for k, v in ipt.items()}, default=str)
        return {
            "SSO":ipt["url_search_string"],
            "IDP:SSOToken": base64.b64encode(
                bytes(jipt.encode("utf-8"))
            ).decode("utf-8")
        }
    except Exception as e:
        traceback.print_exc()
        return {"IDP:SSOToken": "Invalid User"}


# Response Function
def responsedata(status, message, data=None):
    if status:
        return {"status": status, "message": message, "data": data}
    else:
        return {"status": status, "message": message}


# Auth Module


class Authorize(APIView):
    """Login API"""

    def post(self, request):
        if request.data.get("username") and request.data.get("password"):

            key_token = keycloak_auth(username=request.data.get("username"), password=request.data.get("password"))
            # if key_token == "error":
            #     return Response(responsedata(False, "Invalid Details"), status=status.HTTP_400_BAD_REQUEST)
            request.session['raw_password'] = request.data.get("password")
            request.session['token_key'] = key_token
            try:
                params = request.data
                user = User.objects.get(username=params.get("username"))
                if not user.is_active:
                    return Response(
                        responsedata(False, "Your account is disabled"), status=status.HTTP_400_BAD_REQUEST
                    )
                token = create_or_get_token(user)
                request.session['session_tokwn'] = token
                try:
                    user_qs = UserSettings.objects.filter(userName=params.get("username"))
                    user = (user_qs.values("uuid", "role", "sub_role")).first()

                    customers = list(
                        user_qs.annotate(str_id=Cast("customers", output_field=TextField())).values_list(
                            "str_id", flat=True
                        )
                    )
                    annot_kw = dict(
                        assessment=Count('registered_assessments', distinct=True),
                        apprenticeship=Count('registered_apprenticeships', distinct=True),
                    )
                    reg = user_qs.annotate(**annot_kw).values(*annot_kw.keys()).first()
                    response_data = {
                        "token": token,
                        "open_id": key_token,
                        "message": "authorized successfully",
                        "uuid": user["uuid"],
                        "session_data": base64.b64encode(
                            bytes(json.dumps(params).encode('utf-8'))
                        ).decode('utf-8'),
                        "role": user["role"] or 'user',
                        "sub_role": user.get("sub_role", '') or '',
                        "customers": customers,
                        "registered": reg
                    }

                    if user["role"] in ["projectadmin", "user", "customeradmin"]:
                        annotations = dict(
                            company_logo=F("logo"),
                            company_name=F("name"),
                        )
                        values = annotations.keys()
                        if user["role"] in ["projectadmin",]:
                            response_data['project_id'] = user_qs.values_list('allocated_project__uuid', flat=True)[0]
                        # values.extend(['theme__config', 'theme__sub_domain'])
                        qs = Company.objects.filter(uuid=customers[0]).annotate(**annotations)
                        response_data.update(qs.values(*values).first())
                    response_data.update(get_idptoken(**params))
                    request.session['session_data'] = response_data
                    return Response(
                        responsedata(True, "Authorized Successfully", response_data),
                        status=status.HTTP_200_OK,
                    )
                except Exception as e:
                    traceback.print_exc()
                    pass

                response_data = {
                    "token": token,
                    "open_id": key_token,
                    "message": "authorized successfully",
                    "session_data": base64.b64encode(
                        bytes(json.dumps(params).encode('utf-8'))
                    ).decode('utf-8'),
                }
                response_data.update(get_idptoken(**params))
                return Response(
                    responsedata(True, "Authorized Successfully", response_data),
                    status=status.HTTP_200_OK,
                )

            except Exception as e:
                traceback.print_exc()
                return Response(
                    responsedata(False, "Invalid Details"),
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            return Response(responsedata(False, "Details not provided"), status=status.HTTP_400_BAD_REQUEST)


class DisableAccount(APIView):
    """To disable an account"""

    def post(self, request):

        if request.auth is None:
            return Response(responsedata(False, "You are Unauthorized"), status=status.HTTP_400_BAD_REQUEST)

        user = UserSettings.objects.filter(userName=request.user.username).values("role").first()
        if not user["role"] in ['superadmin', 'customeradmin', 'projectadmin']:
            return Response(responsedata(False, "You are Unauthorized"), status=status.HTTP_400_BAD_REQUEST)

        if request.data.get("uuid"):
            user = UserSettings.objects.filter(uuid=request.data.get("uuid")).values("userName").first()
            user = User.objects.get(username=user["userName"])
            if user.is_active:
                user.is_active = False
                msg = "dis"
            else:
                user.is_active = True
                msg = "en"
            user.save()
            return Response(responsedata(True, f"Account {msg}abled"), status=status.HTTP_200_OK)

        else:
            return Response(responsedata(False, "User details not provided"), status=status.HTTP_400_BAD_REQUEST)


class CustomerLogo(APIView):
    """To get logo of a customer"""

    def get(self, request, pk):

        customer_logo = list(Company.objects.filter(uuid=pk).values_list("logo", flat=True))

        return Response(
            responsedata(True, "Customer switched", customer_logo),
            status=status.HTTP_200_OK,
        )


class ForgetPassword(APIView):
    """to send password reset link"""

    def post(self, request):

        if request.data.get("email"):

            email = request.data.get("email")

            if not UserSettings.objects.filter(email=email).exists():
                return Response(responsedata(False, "Email id does not exist"), status=status.HTTP_400_BAD_REQUEST)

            user = User.objects.filter(email=email).values().first()
            password = randomString()

            # Updating password on keycloak
            try:
                keyadmin = keycloak_admin.KeycloakAdmin(
                    server_url=deviare_settings.KEYCLOAK_URL,
                    username=deviare_settings.KEYCLOAK_ADMINUSER,
                    password=deviare_settings.KEYCLOAK_ADMINPASSWORD,
                    realm_name=deviare_settings.KEYCLOAK_REALM,
                )

                user_id_keycloak = keyadmin.get_user_id(user["username"])
                response = keyadmin.set_user_password(user_id=user_id_keycloak, password=password, temporary=False)

            except:
                return Response(responsedata(False, "Something went wrong"), status=status.HTTP_400_BAD_REQUEST)

            # Updating in User table
            auth_user = User.objects.get(username=user["username"])
            auth_user.set_password(password)
            auth_user.save()

            # Updating in User settings
            data = {"password": auth_user.password}
            user_settings = UserSettings.objects.filter(userName=user["username"]).first()
            serializer = UserSettingsSerializer(user_settings, data=data, partial=True)
            if serializer.is_valid(raise_exception=True):

                body1 = "Hi,\n\nYour password has been reset, please find the new login details below\n"
                body2 = "Username : " + user["username"] + "\n" + "Password : " + password + "\n"
                body3 = "Thanks, \n Deviare Support Team"
                send_email(
                    [email],
                    "Deviare Login Details",
                    body1 + body2 + body3,
                )

                serializer.save()
                content = {"user_id": user_settings.user_id_talentlms, "password": password}
                edit_user_profile_on_talent_lms.apply_async(kwargs=content)
                return Response(
                    responsedata(True, "Password sent to registered email address"), status=status.HTTP_200_OK
                )
        else:
            return Response(responsedata(False, "Please provide email"), status=status.HTTP_400_BAD_REQUEST)


class ResetPassword(APIView):
    """to reset password after receiving link for password reset"""

    def post(self, request):

        if request.data.get("token") and request.data.get("password") and request.data.get("username"):
            resetid = request.data.get("token")
            password = request.data.get("password")
            username = request.data.get("username")
            userSettings = UserSettings.objects.get(user__username=username)

            if str(userSettings.reset_id.hex) == resetid:

                # Updating Password on User
                user = User.objects.get(username=username)
                user.set_password(password)
                user.save()

                # Updating password on keycloak
                keyadmin = keycloak_admin.KeycloakAdmin(
                    server_url=deviare_settings.KEYCLOAK_URL,
                    username=deviare_settings.KEYCLOAK_ADMINUSER,
                    password=deviare_settings.KEYCLOAK_ADMINPASSWORD,
                    realm_name=deviare_settings.KEYCLOAK_REALM,
                )

                user_id_keycloak = keyadmin.get_user_id(username)
                response = keyadmin.set_user_password(user_id=user_id_keycloak, password=password, temporary=False)

                # Updating UserSettings
                userSettings.reset_id = uuid.uuid4()
                userSettings.password = user.password
                userSettings.validated = True
                userSettings.save()
                return Response(
                    responsedata(True, "Password reset successfully"),
                    status=status.HTTP_200_OK,
                )

            else:
                return Response(
                    responsedata(False, "Invalid Token"),
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            return Response(
                responsedata(False, "Details not provided"),
                status=status.HTTP_400_BAD_REQUEST,
            )


class MyProfile(APIView):
    """To view details of logged-in user"""

    def get(self, request):

        if request.auth is None:
            return Response(
                responsedata(False, "You are Unauthorized"),
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = UserSettings.objects.filter(userName=request.user.username)
        serializer = UserSettingsSerializer(user, many=True)
        try:
            company_data = []
            for company in (serializer.data)[0]["customers"]:
                company_name = Company.objects.filter(uuid=company).values("name").first()["name"]
                company_data.append({"company_uuid": str(company), "company_name": company_name})
            (serializer.data)[0]["customers"] = company_data
        except:
            pass

        (serializer.data)[0].update(
            {"member_since": (user.values("created_at").first())["created_at"].strftime("%d/%m/%Y")}
        )

        return Response(
            responsedata(True, "User retrieved successfully", serializer.data),
            status=status.HTTP_200_OK,
        )


class EditProfile(APIView):
    """To edit profile of logged-in user"""

    def put(self, request):
        logger.info(request.user)

        if request.auth is None:
            return Response(responsedata(False, "You are Unauthorized"), status=status.HTTP_400_BAD_REQUEST)

        if request.data:
            try:
                data = request.data
                # Field values not to be edited
                if data.get("userName"):
                    del data["userName"]
                if data.get("password"):
                    del data["password"]
                if data.get("email"):
                    email = data.get("email")
                    del data["email"]

                user = UserSettings.objects.filter(userName=request.user.username).first()

                if request.data.get("profile_image") and not request.data.get("profile_image").startswith("http"):

                    try:
                        s3 = boto3.client(
                            "s3",
                            aws_access_key_id=deviare_settings.deviare_config["AWS_S3_ACCESS_KEY"],
                            aws_secret_access_key=deviare_settings.deviare_config["AWS_S3_SECRET_KEY"],
                        )
                    except Exception as e:
                        return Response(
                            responsedata(False, "Can't connect to Database"), status=status.HTTP_400_BAD_REQUEST
                        )

                    try:
                        data_cid, cid_ex = return_image_extenstion(request.data.get("profile_image"))

                        s3.put_object(
                            Body=data_cid,
                            Bucket=deviare_settings.deviare_config["AWS_S3_BUCKET_NAME"],
                            Key="tmp/user/{}.{}".format(email, cid_ex),
                            ACL="public-read",
                        )

                        data.update(
                            {
                                "profile_image": "https://elearn-stat.s3.amazonaws.com/tmp/user/{}.{}".format(
                                    email, cid_ex
                                )
                            }
                        )

                    except Exception as e:
                        return Response(
                            responsedata(False, "Can't Upload Profile Picture"), status=status.HTTP_400_BAD_REQUEST
                        )

                serializer = UserSettingsSerializer(user, data=data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    data = dict(serializer.validated_data)
                    # Call background task on edited field
                    edit_user_profile_on_talent_lms(kwargs=data)
                    return Response(
                        responsedata(True, "Profile updated successfully", serializer.data), status=status.HTTP_200_OK
                    )
            except:
                return Response(responsedata(False, "Something went wrong"), status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(responsedata(False, "Details not provided"), status=status.HTTP_400_BAD_REQUEST)


class ChangePassword(APIView):
    """To update password of logged-in user"""

    def post(self, request):

        if request.auth is None:
            return Response(
                responsedata(False, "You are Unauthorized"),
                status=status.HTTP_400_BAD_REQUEST,
            )

        auth_user = request.user
        if (
            request.data.get("old_password")
            and request.data.get("new_password")
            and request.data.get("confirm_password")
        ):

            if request.data.get("new_password") != request.data.get("confirm_password"):
                return Response(responsedata(False, "Passwords don't Match"), status=status.HTTP_400_BAD_REQUEST)

        else:
            return Response(responsedata(False, "Enough details not provided"), status=status.HTTP_400_BAD_REQUEST)

        key_token = keycloak_auth(username=request.user.username, password=request.data.get("old_password"))
        if key_token == "error":
            return Response(
                responsedata(False, "Old password is incorrect"),
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Updating password on keycloak
        try:
            keyadmin = keycloak_admin.KeycloakAdmin(
                server_url=deviare_settings.KEYCLOAK_URL,
                username=deviare_settings.KEYCLOAK_ADMINUSER,
                password=deviare_settings.KEYCLOAK_ADMINPASSWORD,
                realm_name=deviare_settings.KEYCLOAK_REALM,
            )

            user_id_keycloak = keyadmin.get_user_id(auth_user.username)
            response = keyadmin.set_user_password(
                user_id=user_id_keycloak,
                password=request.data.get("new_password"),
                temporary=False,
            )
        except:
            return Response(
                responsedata(False, "Something went wrong"),
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Updating in User and UserSettings
        user = User.objects.get(username=auth_user.username)
        user.set_password(request.data.get("new_password"))
        user.save()

        data = {"password": user.password}
        user = UserSettings.objects.filter(userName=auth_user.username).first()
        serializer = UserSettingsSerializer(user, data=data, partial=True)
        if serializer.is_valid():
            # TODO updated user password on talentlms
            serializer.save()

        return Response(
            responsedata(True, "Profile updated successfully"),
            status=status.HTTP_200_OK,
        )


class ContactUs(APIView):
    """To contact with deviare support team"""

    def post(self, request):

        data = request.data

        if data.get("name") and data.get("email") and data.get("subject") and data.get("message"):

            try:
                msg_text = render_to_string(
                    "contact.txt",
                    {
                        "name": data.get("name"),
                        "email": data.get("email"),
                        "message": data.get("message"),
                    },
                )
                msg_html = render_to_string(
                    "contact.html",
                    {
                        "name": data.get("name"),
                        "email": data.get("email"),
                        "message": data.get("message"),
                    },
                )

                send_email(
                    ["support@deviare.co.za"],
                    data.get("subject"),
                    msg_text,
                    msg_html,
                )

                return Response(
                    responsedata(True, "Mail sent successfully"),
                    status=status.HTTP_200_OK,
                )
            except:
                return Response(
                    responsedata(False, "Can't send mail"),
                    status=status.HTTP_400_BAD_REQUEST,
                )

        else:
            return Response(
                responsedata(False, "Details missing"),
                status=status.HTTP_400_BAD_REQUEST,
            )


# Administrator APIs


class SuperAdminDashboard(APIView):
    """To get count of customers, projects and users"""

    def get(self, request):
        try:
            data = {}
            user = UserSettings.objects.get(email=request.user.email)
            com_id = [i.hex for i in Project.objects.filter().values_list('company_id', flat=True)]
            data["projects"] = Project.objects.filter(superadmin_id=user).count()
            # data["customers"] =  UserSettings.objects.get(email=request.user.email).customers.all().count()
            data["customers"] =  Company.objects.filter(uuid__in=com_id).count()
            data["users"] = UserSettings.objects.filter(role="user").count()

            return Response(
                responsedata(True, "Dashboard Data fetched successfully", data),
                status=status.HTTP_200_OK,
            )
        except:
            return Response(
                responsedata(False, "Something went wrong"),
                status=status.HTTP_400_BAD_REQUEST,
            )


class SuperAdminList(generics.ListAPIView):
    """To list all superadmins"""

    search_fields = ["firstName", "lastName", "userName", "email"]
    filter_backends = (filters.SearchFilter,)

    def get_queryset(self):
        return UserSettings.objects.filter(role="superadmin").order_by("userName")

    def filter_queryset(self, queryset):
        for backend in list(self.filter_backends):
            queryset = backend().filter_queryset(self.request, queryset, self)
        return queryset

    def get(self, request):

        instance = self.filter_queryset(self.get_queryset())
        admin_list = list(instance.values())

        pagenumber = request.GET.get("page", 1)
        paginator = Paginator(admin_list, 10)

        if int(pagenumber) > paginator.num_pages:
            raise ValidationError("Not enough pages", code=404)
        try:
            previous_page_number = paginator.page(pagenumber).previous_page_number()
        except EmptyPage:
            previous_page_number = None
        try:
            next_page_number = paginator.page(pagenumber).next_page_number()
        except EmptyPage:
            next_page_number = None
        data_to_show = paginator.page(pagenumber).object_list

        return JsonResponse(
            {
                "pagination": {
                    "previous_page": previous_page_number,
                    "is_previous_page": paginator.page(pagenumber).has_previous(),
                    "next_page": next_page_number,
                    "is_next_page": paginator.page(pagenumber).has_next(),
                    "start_index": paginator.page(pagenumber).start_index(),
                    "end_index": paginator.page(pagenumber).end_index(),
                    "total_entries": paginator.count,
                    "total_pages": paginator.num_pages,
                    "page": int(pagenumber),
                },
                "results": data_to_show,
            },
            safe=False,
        )


class SuperAdminCreate(APIView):
    """To create super admin"""

    def post(self, request):
        data = request.data
        data["email"] = data["email"].lower()
        if data["firstName"] and data["lastName"] and data["userName"] and data["email"]:

            if email_check(data["email"]):
                return Response(responsedata(False, "User exists with same email"), status=status.HTTP_400_BAD_REQUEST)

            data["password"] = randomString()

            try:
                # Creating user in keycloak
                keyadmin = keycloak_admin.KeycloakAdmin(
                    server_url=deviare_settings.KEYCLOAK_URL,
                    username=deviare_settings.KEYCLOAK_ADMINUSER,
                    password=deviare_settings.KEYCLOAK_ADMINPASSWORD,
                    realm_name=deviare_settings.KEYCLOAK_REALM,
                )
                # keyadmin = keycloak_admin.KeycloakAdmin(server_url=deviare_settings.KEYCLOAK_URL,username=deviare_settings.KEYCLOAK_ADMINUSER,
                #     password=deviare_settings.KEYCLOAK_ADMINPASSWORD,realm_name=deviare_settings.KEYCLOAK_REALM)

                keyadmin.create_user(
                    {
                        "email": data["email"],
                        "enabled": True,
                        "username": data["userName"],
                        "credentials": [{"value": data["password"], "type": "password"}],
                        "realmRoles": ["user_default"],
                    }
                )

                # Creating user in User Table
                user = User(email=data["email"], username=data["userName"])
                user.set_password(data["password"])
                user.save()

                # Adding entry in UserSettings Table
                if request.data.get("profile_image") and not request.data.get("profile_image").startswith("http"):

                    try:
                        s3 = boto3.client(
                            "s3",
                            aws_access_key_id=deviare_settings.deviare_config["AWS_S3_ACCESS_KEY"],
                            aws_secret_access_key=deviare_settings.deviare_config["AWS_S3_SECRET_KEY"],
                        )
                    except Exception as e:
                        return Response(
                            responsedata(False, "Can't connect to Database"), status=status.HTTP_400_BAD_REQUEST
                        )

                    try:
                        data_cid, cid_ex = return_image_extenstion(request.data.get("profile_image"))

                        s3.put_object(
                            Body=data_cid,
                            Bucket=deviare_settings.deviare_config["AWS_S3_BUCKET_NAME"],
                            Key="tmp/user/{}.{}".format(request.data.get("email"), cid_ex),
                            ACL="public-read",
                        )

                        data.update(
                            {
                                "profile_image": "https://elearn-stat.s3.amazonaws.com/tmp/user/{}.{}".format(
                                    request.data.get("email"), cid_ex
                                )
                            }
                        )

                    except Exception as e:
                        return Response(
                            responsedata(False, "Can't Upload Profile Picture"), status=status.HTTP_400_BAD_REQUEST
                        )

                data["user"] = user.id
                data["role"] = "superadmin"
                temp_password = data["password"]
                data["password"] = user.password
                serializer = UserSettingsSerializer(data=data)
                if serializer.is_valid(raise_exception=True):
                    body1 = "Hi,\n\n\n\nYour account has been created, please find the login details below\n\n"
                    body2 = "Username : " + data["userName"] + "\n\n" + "Password : " + temp_password +  "\n\nURL : " + deviare_settings.URL +"/\n\n"
                    body3 = "Thanks, \n\n Deviare Support Team"
                    send_email([data["email"]], "Deviare Login Details", body1 + body2 + body3)
                    serializer.save()
                    return Response(
                        responsedata(True, "Super Admin Added", serializer.data), status=status.HTTP_200_OK
                    )
            except:
                return Response(responsedata(False, "User already exists"), status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(responsedata(False, "Details Missing"), status=status.HTTP_400_BAD_REQUEST)


class SuperAdminDetail(APIView):
    """To retrieve and update super admin"""

    def get(self, request, pk):

        user = UserSettings.objects.filter(uuid=pk)
        serializer = UserSettingsSerializer(user, many=True)

        user = user.values("userName").first()
        user = User.objects.get(username=user["userName"])
        data = {"is_active": user.is_active}
        data.update(serializer.data[0])

        if not type(data) is list:
            data = [data]

        return Response(responsedata(True, "Super Admin retrieved successfully", data), status=status.HTTP_200_OK)

    def put(self, request, pk):

        if request.auth is None:
            return Response(responsedata(False, "You are Unauthorized"), status=status.HTTP_400_BAD_REQUEST)

        authuser = UserSettings.objects.filter(userName=request.user.username).values().first()

        if not (authuser["role"] == "superadmin"):
            return Response(responsedata(False, "You are not authorized"), status=status.HTTP_400_BAD_REQUEST)

        if request.data.get("userName"):
            del request.data["userName"]
        if request.data.get("email"):
            email = request.data.get("email")
            del request.data["email"]

        if request.data.get("profile_image") and not request.data.get("profile_image").startswith("http"):

            try:
                s3 = boto3.client(
                    "s3",
                    aws_access_key_id=deviare_settings.deviare_config["AWS_S3_ACCESS_KEY"],
                    aws_secret_access_key=deviare_settings.deviare_config["AWS_S3_SECRET_KEY"],
                )
            except Exception as e:
                return Response(responsedata(False, "Can't connect to Database"), status=status.HTTP_400_BAD_REQUEST)

            try:
                data_cid, cid_ex = return_image_extenstion(request.data.get("profile_image"))

                s3.put_object(
                    Body=data_cid,
                    Bucket=deviare_settings.deviare_config["AWS_S3_BUCKET_NAME"],
                    Key="tmp/user/{}.{}".format(email, cid_ex),
                    ACL="public-read",
                )

                request.data.update(
                    {"profile_image": "https://elearn-stat.s3.amazonaws.com/tmp/user/{}.{}".format(email, cid_ex)}
                )

            except Exception as e:
                return Response(
                    responsedata(False, "Can't Upload Profile Picture"), status=status.HTTP_400_BAD_REQUEST
                )

        change_password = False
        if request.data.get("password"):

            user = UserSettings.objects.filter(uuid=pk).values().first()
            password = request.data.get("password")

            # Updating password on keycloak
            try:
                keyadmin = keycloak_admin.KeycloakAdmin(
                    server_url=deviare_settings.KEYCLOAK_URL,
                    username=deviare_settings.KEYCLOAK_ADMINUSER,
                    password=deviare_settings.KEYCLOAK_ADMINPASSWORD,
                    realm_name=deviare_settings.KEYCLOAK_REALM,
                )

                user_id_keycloak = keyadmin.get_user_id(user["userName"])
                response = keyadmin.set_user_password(user_id=user_id_keycloak, password=password, temporary=False)

            except:
                return Response(responsedata(False, "Something went wrong"), status=status.HTTP_400_BAD_REQUEST)

            # Updating in User table
            auth_user = User.objects.get(username=user["userName"])
            auth_user.set_password(password)
            auth_user.save()

            request.data["password"] = auth_user.password
            request.data["last_pw_update_date"] = datetime.datetime.now()
            request.data["last_pw_update_by"] = authuser["email"]
            change_password = True

        admin = UserSettings.objects.filter(uuid=pk).first()

        serializer = UserSettingsSerializer(admin, data=request.data, partial=True)
        if serializer.is_valid(raise_exception=True):
            if change_password:
                body1 = "Hi,\n\n\n\nYour password has been reset, please find the new login details below\n\n"
                body2 = "Username : " + user["userName"] + "\n\n" + "Password : " + password +  "\n\nURL : " + deviare_settings.URL +"/\n\n"
                body3 = "Thanks, \n\n Deviare Support Team"
                send_email([email], "Deviare Login Details", body1 + body2 + body3)

            serializer.save()
            return Response(
                responsedata(True, "Administrator updated successfully", serializer.data), status=status.HTTP_200_OK
            )
        return Response(
            responsedata(False, "Something went wrong", serializer.errors), status=status.HTTP_400_BAD_REQUEST
        )

    def delete(self, request, pk):

        userName = UserSettings.objects.filter(uuid=pk).values("userName").first()["userName"]

        # Deleting from User
        try:
            user = User.objects.get(username=userName)
            user.delete()
        except:
            pass

        # Deleting from UserSettings
        try:
            userSettings = UserSettings.objects.filter(uuid=pk)
            userSettings.delete()
        except:
            pass

        # Deleting from Keycloak
        keyadmin = keycloak_admin.KeycloakAdmin(
            server_url=deviare_settings.KEYCLOAK_URL,
            username=deviare_settings.KEYCLOAK_ADMINUSER,
            password=deviare_settings.KEYCLOAK_ADMINPASSWORD,
            realm_name=deviare_settings.KEYCLOAK_REALM,
        )

        try:
            keyid = keyadmin.get_user_id(userName)
            keyadmin.delete_user(user_id=keyid)
        except:
            pass

        return Response(responsedata(True, "User Deleted"), status=status.HTTP_200_OK)


# Customer APIs

class GCologistDashboard(APIView):
    """To get details for customer admin dashboard"""
    @data_response
    def get(self, request):

        if request.auth is None:
            return Response(
                responsedata(False, "You are Unauthorized"),
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = request.userAuth
        filter_kwargs = {
            'gcologist': request.userAuth
            # 'project_admin' if user.role == 'projectadmin' else 'company_admin_id': user.uuid
        }

        def grf(x):
            if x == 10:
                return 'In Progress'
            if x == 9:
                return 'Completed Review'
            if x >= 6:
                return 'Pending Review'
            return 'Pending'

        qs = GCIndexAssessment.objects.filter(**filter_kwargs)
        ser = pd.DataFrame(GCIndexAssessmentSerializer(qs, many=True, fields=['state_id']).data)
        ser['title'] = ''
        ser['title'] = ser.state_id.apply(grf)
        ser = ser.rename(index=str, columns={'state_id': 'value'})

        data = ser.groupby('title', as_index=False).count().to_dict(orient='records')
        # data.update(open_close)
        return Response(
            responsedata(True, "GCologist dashboard", data),
            status=status.HTTP_200_OK,
        )


class CustomerAdminDashboard(APIView):
    """To get details for customer admin dashboard"""
    @data_response
    def get(self, request):

        if request.auth is None:
            return Response(
                responsedata(False, "You are Unauthorized"),
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = request.userAuth
        filter_kwargs = {
            'project_admin' if user.role == 'projectadmin' else 'company_admin_id': user.uuid
        }
        projects = Project.objects.filter(**filter_kwargs)
        ser = ProjectDashboardSerializer(projects, fields=[
            'status',
            'total_course_licenses',
            'consumed_course_licenses',
            'course_completed_count',
            'total_assessment_licenses',
            'consumed_assessment_licenses',
            'total_licenses',
            'total_users',
            'user_count',
            'total_license_count'], many=True).df
        # open_close = ser[['status']].groupby('status').count().to_dict(orient='records')

        data = ser.sum().to_dict()
        # data.update(open_close)
        return Response(
            responsedata(True, "Customer admin dashboard", data),
            status=status.HTTP_200_OK,
        )


class CustomerList(generics.ListAPIView):
    """To list customers"""

    search_fields = ["name", "country", "address"]
    filter_backends = (filters.SearchFilter,)

    def get_queryset(self):
        return Company.objects.all().order_by("name")

    def filter_queryset(self, queryset):
        for backend in list(self.filter_backends):
            queryset = backend().filter_queryset(self.request, queryset, self)
        return queryset

    def get(self, request):

        instance = self.filter_queryset(self.get_queryset())
        companies_list = list(instance.values())

        pagenumber = request.GET.get("page", 1)
        paginator = Paginator(companies_list, 10)

        if int(pagenumber) > paginator.num_pages:
            raise ValidationError("Not enough pages", code=404)
        try:
            previous_page_number = paginator.page(pagenumber).previous_page_number()
        except EmptyPage:
            previous_page_number = None
        try:
            next_page_number = paginator.page(pagenumber).next_page_number()
        except EmptyPage:
            next_page_number = None
        data_to_show = paginator.page(pagenumber).object_list

        return JsonResponse(
            {
                "pagination": {
                    "previous_page": previous_page_number,
                    "is_previous_page": paginator.page(pagenumber).has_previous(),
                    "next_page": next_page_number,
                    "is_next_page": paginator.page(pagenumber).has_next(),
                    "start_index": paginator.page(pagenumber).start_index(),
                    "end_index": paginator.page(pagenumber).end_index(),
                    "total_entries": paginator.count,
                    "total_pages": paginator.num_pages,
                    "page": int(pagenumber),
                },
                "results": data_to_show,
            },
            safe=False,
        )


class CustomerCreate(APIView):
    """To create a customer"""

    def post(self, request):

        if request.data:

            if Company.objects.filter(name=request.data.get("name")).exists():
                return Response(
                    responsedata(False, "Customer name already exists"), status=status.HTTP_400_BAD_REQUEST
                )

            if request.data.get("logo") and not request.data.get("logo").startswith("http"):

                try:
                    s3 = boto3.client(
                        "s3",
                        aws_access_key_id=deviare_settings.deviare_config["AWS_S3_ACCESS_KEY"],
                        aws_secret_access_key=deviare_settings.deviare_config["AWS_S3_SECRET_KEY"],
                    )
                except Exception as e:
                    return Response(
                        responsedata(False, "Can't connect to Database"), status=status.HTTP_400_BAD_REQUEST
                    )

                try:
                    data_cid, cid_ex = return_image_extenstion(request.data.get("logo"))

                    filename = (request.data.get("name")).replace(" ", "%20")

                    s3.put_object(
                        Body=data_cid,
                        Bucket=deviare_settings.deviare_config["AWS_S3_BUCKET_NAME"],
                        Key="tmp/customer/{}.{}".format(filename, cid_ex),
                        ACL="public-read",
                    )

                    request.data.update(
                        {"logo": "https://elearn-stat.s3.amazonaws.com/tmp/customer/{}.{}".format(filename, cid_ex)}
                    )

                except Exception as e:
                    return Response(
                        responsedata(False, "Can't Upload Profile Picture"), status=status.HTTP_400_BAD_REQUEST
                    )

            serializer = CompanySerializer(data=request.data)
            if serializer.is_valid(raise_exception=True):
                serializer.save()
                return Response(
                    responsedata(True, "Customer Created successfully", serializer.data), status=status.HTTP_200_OK
                )
        else:
            return Response(responsedata(False, "Details not provided"), status=status.HTTP_400_BAD_REQUEST)


class CustomerDetail(APIView):
    """Retrieve, update or delete a customer"""

    def get(self, request, pk):
        customer = Company.objects.filter(uuid=pk)
        serializer = CompanySerializer(customer, many=True)
        return Response(
            responsedata(True, "Customer retrieved successfully", serializer.data),
            status=status.HTTP_200_OK,
        )

    def put(self, request, pk):

        customer = Company.objects.filter(uuid=pk).first()

        if request.data.get("logo") and not request.data.get("logo").startswith("http"):

            try:
                s3 = boto3.client(
                    "s3",
                    aws_access_key_id=deviare_settings.deviare_config["AWS_S3_ACCESS_KEY"],
                    aws_secret_access_key=deviare_settings.deviare_config["AWS_S3_SECRET_KEY"],
                )
            except Exception as e:
                return Response(responsedata(False, "Can't connect to Database"), status=status.HTTP_400_BAD_REQUEST)

            try:
                data_cid, cid_ex = return_image_extenstion(request.data.get("logo"))
                filename = (request.data.get("name")).replace(" ", "%20")

                s3.put_object(
                    Body=data_cid,
                    Bucket=deviare_settings.deviare_config["AWS_S3_BUCKET_NAME"],
                    Key="tmp/customer/{}.{}".format(filename, cid_ex),
                    ACL="public-read",
                )

                request.data.update(
                    {"logo": "https://elearn-stat.s3.amazonaws.com/tmp/customer/{}.{}".format(filename, cid_ex)}
                )

            except Exception as e:
                return Response(
                    responsedata(False, "Can't Upload Profile Picture"), status=status.HTTP_400_BAD_REQUEST
                )

        if request.data.get("name"):
            if Company.objects.filter(name=request.data.get("name")).exists() and customer.name != request.data.get(
                "name"
            ):
                return Response(
                    responsedata(False, "Customer name already exists"), status=status.HTTP_400_BAD_REQUEST
                )

        serializer = CompanySerializer(customer, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(
                responsedata(True, "Customer updated successfully", serializer.data), status=status.HTTP_200_OK
            )
        return Response(
            responsedata(False, "Details not provided", serializer.errors), status=status.HTTP_400_BAD_REQUEST
        )

    def delete(self, request, pk):
        customer = Company.objects.filter(uuid=pk)
        customer.delete()
        return Response(responsedata(True, "Customer Deleted"), status=status.HTTP_200_OK)


# User APIs


class UserDashboard(APIView):
    """To get details for user dashboard"""

    def get(self, request):

        if request.auth is None:
            return Response(responsedata(False, "You are Unauthorized"), status=status.HTTP_400_BAD_REQUEST)

        data = {
            "courses": CourseLicenseUser.objects.filter(user_id__userName=request.user.username).count(),
            "completed_courses": CourseLicenseUser.objects.filter(
                user_id__userName=request.user.username, course_completion=float(100)
            ).count(),
        }

        return Response(
            responsedata(True, "User details retrieved successfully", data),
            status=status.HTTP_200_OK,
        )


class UserList(generics.ListAPIView):
    """To list users, search and pagination implemented"""

    search_fields = ["firstName", "lastName", "userName", "email"]
    filter_backends = (filters.SearchFilter,)

    def get_queryset(self, role, user):

        if role == "superadmin":
            return UserSettings.objects.exclude(role='superadmin').order_by("userName")

        elif role == "projectadmin":
            company_id = list(UserSettings.objects.filter(uuid=user).values_list("allocated_project__company_id", flat=True))
            return UserSettings.objects.filter(role__in=["gcologist", "user"], customers__in=company_id).order_by("userName")
        elif role == "customeradmin":
            company_id = list(UserSettings.objects.filter(uuid=user).values_list("customers", flat=True))
            return UserSettings.objects.filter(role__in=["projectadmin", "gcologist", "user"], customers__in=company_id).order_by(
                "userName")

        else:
            return Response(
                responsedata(False, "You are Unauthorized"),
                status=status.HTTP_400_BAD_REQUEST,
            )

    def filter_queryset(self, queryset):
        for backend in list(self.filter_backends):
            queryset = backend().filter_queryset(self.request, queryset, self)
        return queryset

    def get(self, request):

        if request.auth is None:
            return Response(
                responsedata(False, "You are Unauthorized"),
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = UserSettings.objects.filter(userName=request.user.username).values().first()
        instance = self.filter_queryset(self.get_queryset(user["role"], user["uuid"]))
        user_list = list(instance.values())

        pagenumber = request.GET.get("page", 1)
        paginator = Paginator(user_list, 10)

        if int(pagenumber) > paginator.num_pages:
            raise ValidationError("Not enough pages", code=404)
        try:
            previous_page_number = paginator.page(pagenumber).previous_page_number()
        except EmptyPage:
            previous_page_number = None
        try:
            next_page_number = paginator.page(pagenumber).next_page_number()
        except EmptyPage:
            next_page_number = None
        data_to_show = paginator.page(pagenumber).object_list

        return JsonResponse(
            {
                "pagination": {
                    "previous_page": previous_page_number,
                    "is_previous_page": paginator.page(pagenumber).has_previous(),
                    "next_page": next_page_number,
                    "is_next_page": paginator.page(pagenumber).has_next(),
                    "start_index": paginator.page(pagenumber).start_index(),
                    "end_index": paginator.page(pagenumber).end_index(),
                    "total_entries": paginator.count,
                    "total_pages": paginator.num_pages,
                    "page": int(pagenumber),
                },
                "results": data_to_show,
            },
            safe=False,
        )


class UserCreate(APIView):
    """To create user or customer admin"""

    def post(self, request):

        logger.info(request.data)
        if request.auth is None:
            return Response(responsedata(False, "You are Unauthorized"), status=status.HTTP_400_BAD_REQUEST)

        data = request.data
        if not data.get("customers"):
            return Response(responsedata(False, "Please select a Company"), status=status.HTTP_400_BAD_REQUEST)

        user = UserSettings.objects.filter(userName=request.user.username).values().first()

        if user["role"] == "superadmin":

            if not (data.get("role") in ["projectadmin", "user", "customeradmin", "gcologist"]):
                return Response(
                    responsedata(False, "Role Missing or invalid role"), status=status.HTTP_400_BAD_REQUEST
                )

            if data.get("role") == "customeradmin" and len(data["customers"]) > 1:
                return Response(
                    responsedata(False, "Select one company only for customer admin"),
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for customer in data["customers"]:
                if Company.objects.filter(uuid=customer) is None:
                    return Response(responsedata(False, "Invalid Company"), status=status.HTTP_400_BAD_REQUEST)
        customers = data.get("customers", False)
        if user["role"] == "projectadmin":
            if not data.get("role") in ["user", ]:
                return Response(responsedata(False, "Invalid Role"), status=status.HTTP_400_BAD_REQUEST)

        if user["role"] == "customeradmin":
            if not data.get("role") in ["projectadmin", "user", "gcologist"]:
                return Response(responsedata(False, "Invalid Role"), status=status.HTTP_400_BAD_REQUEST)
        if user["role"] in ['projectadmin', 'customeradmin']:
            if customers is not False and len(customers) > 1:
                return Response(responsedata(False, "Select one company only"), status=status.HTTP_400_BAD_REQUEST)
            if data.get("role") in ["projectadmin", "user"]:
                user_customers = str(
                    list(UserSettings.objects.filter(uuid=user["uuid"]).values_list("customers", flat=True))[0]
                ).replace("-", "")
                if user_customers != (data["customers"][0]).replace("-", ""):
                    return Response(responsedata(False, "Invalid Company"), status=status.HTTP_400_BAD_REQUEST)

        if data["firstName"] and data["lastName"] and data["userName"] and data["email"]:
            data["email"] = data["email"].lower()
            # for existing users
            if email_check(data["email"]):
                if customers is not False:
                    for customer in data["customers"]:
                        if UserSettings.objects.filter(email=data["email"], customers=customer).exists():
                            return Response(
                                responsedata(False, "User exists with same email for this customer"),
                                status=status.HTTP_400_BAD_REQUEST,
                            )
                        else:
                            user = UserSettings.objects.get(email=data["email"])
                            user.customers.add(customer)
                return Response(responsedata(True, "User added"), status=status.HTTP_200_OK)

            # for new users
            data["password"] = randomString()

            try:

                from main.fixer import find_keycloak_user, delete_keycloak_user
                exists = find_keycloak_user(email=data["email"])
                if exists is not False:
                    delete_keycloak_user(email=data["email"])
                    # Creating user in keycloak
                keyadmin = keycloak_admin.KeycloakAdmin(
                        server_url=deviare_settings.KEYCLOAK_URL,
                        username=deviare_settings.KEYCLOAK_ADMINUSER,
                        password=deviare_settings.KEYCLOAK_ADMINPASSWORD,
                        realm_name=deviare_settings.KEYCLOAK_REALM,
                    )
                logger.info("creating user with this credentials")
                logger.info(data)
                keyadmin.create_user(
                    {
                        "email": data["email"],
                        "enabled": True,
                        "username": data["userName"],
                        "credentials": [{"value": data["password"], "type": "password"}],
                        "realmRoles": ["user_default"],
                    }
                )
            except Exception as e:
                logger.exception(e)
            try:
                # Creating user in User Table
                user, c = User.objects.update_or_create(email=data["email"], username=data["userName"])
                user.set_password(data["password"])
                user.save()
                user.refresh_from_db()
                # Adding entry in UserSettings Table
                data["user"] = user.id

                if request.data.get("profile_image"):

                    try:
                        s3 = boto3.client(
                            "s3",
                            aws_access_key_id=deviare_settings.deviare_config["AWS_S3_ACCESS_KEY"],
                            aws_secret_access_key=deviare_settings.deviare_config["AWS_S3_SECRET_KEY"],
                        )
                    except Exception as e:
                        return Response(
                            responsedata(False, "Can't connect to Database"), status=status.HTTP_400_BAD_REQUEST
                        )

                    try:
                        data_cid, cid_ex = return_image_extenstion(request.data.get("profile_image"))

                        s3.put_object(
                            Body=data_cid,
                            Bucket=deviare_settings.deviare_config["AWS_S3_BUCKET_NAME"],
                            Key="tmp/user/{}.{}".format(request.data.get("email"), cid_ex),
                            ACL="public-read",
                        )

                        data.update(
                            {
                                "profile_image": "https://elearn-stat.s3.amazonaws.com/tmp/user/{}.{}".format(
                                    request.data.get("email"), cid_ex
                                )
                            }
                        )

                    except Exception as e:
                        return Response(
                            responsedata(False, "Can't Upload Profile Picture"), status=status.HTTP_400_BAD_REQUEST
                        )
            except Exception as e:
                logger.exception(e)
            try:
                # update UserSettings
                user_qs = UserSettings.objects.filter(email=data["email"])
                temp_password = data["password"]
                data["password"] = user.password
                # if data.get('user', None):
                #     logger.info("Load User ID and PWD")
                #     pk, pwd = User.objects.filter(email=data["email"]).values_list(
                #         'pk', 'password').first()
                #     data['user'] = pk
                #     data["password"] = pwd
                if user_qs.exists():
                    user_obj = user_qs.first()
                    data.pop('password')
                    user_obj.set_password(temp_password)
                    user_obj.save()
                    serializer = UserSettingsSerializer(user_obj, data=data)
                else:
                    serializer = UserSettingsSerializer(data=data)
                if serializer.is_valid(raise_exception=True):

                    body1 = "Hi,\n\n\n\nYour account has been created, please find the login details below\n\n"
                    body2 = "Username : " + data["userName"] + "\n\n" + "Password : " + temp_password + "\n\nURL : " + deviare_settings.URL +"/\n\n"
                    body3 = "Thanks, \n\n Deviare Support Team"
                    logger.info('BEFORE EMAIL')
                    send_email([data["email"]],"Deviare Login Details",body1 + body2 + body3,)
                    logger.info('After EMAIL')
                    # call background task to create user on talentlms platform
                    # we need the unencrypted password
                    data["unencrypted_password"] = temp_password
                    logger.info(data)
                    add_user_to_talent_lms.apply_async(kwargs=data)

                    serializer.save()

                    return Response(responsedata(True, "User Added", serializer.data), status=status.HTTP_200_OK)
            except Exception as exc:
                logger.exception(exc)
                return Response(responsedata(False, "User already exists"), status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(responsedata(False, "Details Missing"), status=status.HTTP_400_BAD_REQUEST)


class UserDetail(APIView):
    """To retrieve, update or delete user or customer admin"""

    def get(self, request, pk):
        user = UserSettings.objects.filter(uuid=pk)
        serializer = UserSettingsSerializer(user, many=True)

        user = user.values("userName").first()
        user = User.objects.get(username=user["userName"])
        data = {"is_active": user.is_active}
        data.update(serializer.data[0])

        if not type(data) is list:
            data = [data]

        return Response(responsedata(True, "User retrieved successfully", data), status=status.HTTP_200_OK)

    def put(self, request, pk):

        if request.auth is None:
            return Response(responsedata(False, "You are Unauthorized"), status=status.HTTP_400_BAD_REQUEST)
  
        authuser = UserSettings.objects.filter(userName=request.user.username).values().first()

        if not authuser["role"] in ["projectadmin", "superadmin", "customeradmin"]:
            return Response(responsedata(False, "You are not authorized"), status=status.HTTP_400_BAD_REQUEST)

        if request.data.get("userName"):
            del request.data["userName"]
        if request.data.get("email"):
            email = request.data.get("email")
            del request.data["email"]


        try:
            if request.data["role"] == "customeradmin" and len(request.data["customers"]) > 1:
                return Response(
                    responsedata(False, "Select one company only for customer admin"),
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except KeyError as e:
            logger.exception(e)
            pass

        if request.data.get("profile_image") and not request.data.get("profile_image").startswith("http"):

            try:
                s3 = boto3.client(
                    "s3",
                    aws_access_key_id=deviare_settings.deviare_config["AWS_S3_ACCESS_KEY"],
                    aws_secret_access_key=deviare_settings.deviare_config["AWS_S3_SECRET_KEY"],
                )

            except Exception as e:
                logger.exception(e)
                return Response(responsedata(False, "Can't connect to Database"), status=status.HTTP_400_BAD_REQUEST)

            try:
                data_cid, cid_ex = return_image_extenstion(request.data.get("profile_image"))

                s3.put_object(
                    Body=data_cid,
                    Bucket=deviare_settings.deviare_config["AWS_S3_BUCKET_NAME"],
                    Key="tmp/user/{}.{}".format(email, cid_ex),
                    ACL="public-read",
                )

                request.data.update(
                    {"profile_image": "https://elearn-stat.s3.amazonaws.com/tmp/user/{}.{}".format(email, cid_ex)}
                )

            except Exception as e:
                logger.exception(e)
                return Response(
                    responsedata(False, "Can't Upload Profile Picture"), status=status.HTTP_400_BAD_REQUEST
                )

        change_password = False
        if request.data.get("password"):

            user = UserSettings.objects.filter(uuid=pk).values().first()
            password = request.data.get("password")

            # Updating password on keycloak
            try:
                keyadmin = keycloak_admin.KeycloakAdmin(
                    server_url=deviare_settings.KEYCLOAK_URL,
                    username=deviare_settings.KEYCLOAK_ADMINUSER,
                    password=deviare_settings.KEYCLOAK_ADMINPASSWORD,
                    realm_name=deviare_settings.KEYCLOAK_REALM,
                )

                user_id_keycloak = keyadmin.get_user_id(user["userName"])
                response = keyadmin.set_user_password(user_id=user_id_keycloak, password=password, temporary=False)
                print(response)
            except Exception as e:
                logger.exception(e)
                from main.fixer import create_keycloak_user
                try:
                    create_keycloak_user(username=user.get("userName"), password=password, keyadmin=keyadmin,
                                         **{k: user.get(k) for k in ["firstName", "email", "lastName"]})
                except Exception as e:
                    logger.exception(e)
                    return Response(responsedata(False, "Something went wrong"), status=status.HTTP_400_BAD_REQUEST)

            # Updating in User table
            auth_user = User.objects.get(username=user["userName"])
            auth_user.set_password(password)
            auth_user.save()

            request.data["password"] = auth_user.password
            request.data["last_pw_update_date"] = datetime.datetime.now()
            request.data["last_pw_update_by"] = authuser["email"]

            change_password = True

        user_setting = UserSettings.objects.filter(uuid=pk).first()
        print(request.data)
        if request.data.get('is_active', None) is not None:
            us = User.objects.filter(settings__uuid=pk).first()
            us.is_active = request.data.get('is_active')
            user = UserSettings.objects.get(email=us.email)
            action_user = UserSettings.objects.get(email=request.user.email)
            us.save()
            action_id = user.user_id_talentlms
            user_action_id = action_user.user_id_talentlms
            if request.data.get('is_active') == False and action_id != None:
                change_user_status_to_inactive(user_id=action_id)
            if request.data.get('is_active') == True and action_id != None:
                change_user_status_to_active(user_id=action_id)
        serializer = UserSettingsSerializer(user_setting, data=request.data, partial=True)
        if serializer.is_valid(raise_exception=True):

            if change_password:
                body = f"""Hi,
                \n\nYour password has been reset, please find the new login details below\n
                Username :  {user["userName"]} \n
                Password : {password} \n"
                Thanks, \n 
                Deviare Support Team"""
                send_email([email], "Deviare Login Details", body)

            serializer.save()

            return Response(
                responsedata(True, "User updated successfully", serializer.data), status=status.HTTP_200_OK
            )
        return Response(
            responsedata(False, "Details not provided", serializer.errors), status=status.HTTP_400_BAD_REQUEST
        )

    def delete(self, request, pk):

        userName = UserSettings.objects.filter(uuid=pk).values("userName").first()["userName"]

        # Deleting from User
        try:
            user = User.objects.get(username=userName)
            user.delete()
        except:
            pass

        # Deleting from UserSettings
        try:
            userSettings = UserSettings.objects.filter(uuid=pk)
            userSettings.delete()
        except:
            pass

        # Deleting from Keycloak
        keyadmin = keycloak_admin.KeycloakAdmin(
            server_url=deviare_settings.KEYCLOAK_URL,
            username=deviare_settings.KEYCLOAK_ADMINUSER,
            password=deviare_settings.KEYCLOAK_ADMINPASSWORD,
            realm_name=deviare_settings.KEYCLOAK_REALM,
        )

        try:
            keyid = keyadmin.get_user_id(userName)
            keyadmin.delete_user(user_id=keyid)
        except:
            pass

        return Response(responsedata(True, "User Deleted"), status=status.HTTP_200_OK)


# Bulk Upload Users


class BulkUploadTemplate(APIView):
    """To get excel template for bulk upload of users"""

    def get(self, request):

        if request.auth is None:
            return Response(
                responsedata(False, "You are Unauthorized"),
                status=status.HTTP_400_BAD_REQUEST,
            )
       
        user = request.user
        user = UserSettings.objects.filter(userName=user.username).values().first()

        # Creating WorkBook and WorkSheet
        wb = Workbook()
        ws = wb.create_sheet("Main Sheet", 0)

        # Filling the top row with Headings
        headings = ["First name", "Surname", "Username", "Email", "Customer", "Role", "Project"]

        for i in range(len(headings)):
            ws.cell(row=1, column=i + 1).value = headings[i]

        # Column width
        columns = ["A", "B", "C", "D", "E", "F","G"]
        for column in columns:
            ws.column_dimensions[column].width = 30

        # Styling Cells
        for i in range(len(headings)):
            ws.cell(row=1, column=i + 1).font = Font(bold=True)

        # Applying Validation to Role Column
        if user["role"] == "superadmin":
            role_val = DataValidation(type="list", formula1='"user,customeradmin,projectadmin"')
            ws.add_data_validation(role_val)
        

        elif user["role"] in ["projectadmin", "customeradmin"]:
            role_val = DataValidation(type="list", formula1='"user,projectadmin"')
            ws.add_data_validation(role_val)

        else:
            return Response(
                responsedata(False, "You are Unauthorized"),
                status=status.HTTP_400_BAD_REQUEST,
            )

        for i in range(2, 501):
            cell = "F" + str(i)
            role_val.add(ws[cell])

        # Applying Validation to Customers
      
        if user["role"] == "superadmin":

            # Fetching Customers
            customers = Company.objects.values("name").order_by("-updated_at")
            customers_list = [c["name"] for c in customers]
            # customers_list = [c.encode('utf-8') for c in customers_list]
            # customers_list = ",".join(customers_list)
            # customers_list = '"{}"'.format(customers_list)

            # Filling Customers
            for i in range(len(customers_list)):
                ws.cell(row=i + 501, column=54).value = customers_list[i]

            # Applying Validation to Customers Column
            template_range_T = "=$BB501:$BB" + str(len(customers_list) + 500)
            customer_val = DataValidation(type="list", formula1=template_range_T)
            ws.add_data_validation(customer_val)
            customers = Company.objects.values("name").order_by("-updated_at")
            customers_list = [c["name"] for c in customers]
            # customers_list = [c.encode('utf-8') for c in customers_list]
            # customers_list = ",".join(customers_list)
            # customers_list = '"{}"'.format(customers_list)

            # Filling Project
            project = Project.objects.values("project_name").order_by("-updated_at")
            project_list = [c["project_name"] for c in project]
            for i in range(len(project_list)):
                ws.cell(row=i + 501, column=54).value = project_list[i]

            # Applying Validation to Customers Column
            template_range_T = "=$BB501:$BB" + str(len(project) + 500)
            project_val = DataValidation(type="list", formula1=template_range_T)
            ws.add_data_validation(project_val)
           
            ws.column_dimensions["BB"].hidden = True
            project = Project.objects.all()
            project_list = '"{}"'.format(",".join(project.values_list('project_name', flat=True)))

            project_val = DataValidation(type="list", formula1=project_list)
            ws.add_data_validation(project_val)

        elif user["role"] == "customeradmin":

            # Fetching Customers
            customer = list(
                UserSettings.objects.filter(userName=user["userName"]).values_list("customers", flat=True)
            )[0]
            customer = Company.objects.filter(uuid=customer)
            customers_list = '"{}"'.format(",".join(customer.values_list('name', flat=True)))
            customer_val = DataValidation(type="list", formula1=customers_list)
            ws.add_data_validation(customer_val)

            project = Project.objects.filter(company_id__uuid__in=customer)
            project_list = '"{}"'.format(",".join(project.values_list('project_name', flat=True)))

            project_val = DataValidation(type="list", formula1=project_list)
            ws.add_data_validation(project_val)
            

        for i in range(2, 501):
            cell = "E" + str(i)
            customer_val.add(ws[cell])
       
 
        for i in range(2, 501):
            cell = "G" + str(i)
            project_val.add(ws[cell])

        # Locking Sheet
        ws.protection.sheet = False

        # Unlocking Editable Cells
        for r in range(2, 501):
            for c in range(ord("A"), ord("G")):
                c = chr(c)
                cell = str(c) + str(r)
                ws[cell].protection = Protection(locked=False)

        # Saving Workbook
        wb.save("bulkupload.xlsx")

        # S3 Upload
        try:
            s3 = boto3.client(
                "s3",
                aws_access_key_id=deviare_settings.deviare_config["AWS_S3_ACCESS_KEY"],
                aws_secret_access_key=deviare_settings.deviare_config["AWS_S3_SECRET_KEY"],
            )
        except Exception as e:
            return Response(
                responsedata(False, "Can't connect to Database"),
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            with open("bulkupload.xlsx", "rb") as f:
                data = f.read()

            data1 = base64.b64encode(data)
            data2 = base64.b64decode(data1)
            s3.put_object(
                Body=data2,
                Bucket="elearn-stat",
                Key="tmp/bulkupload.xlsx",
                ACL="public-read",
            )

            url = "https://elearn-stat.s3.amazonaws.com/tmp/bulkupload.xlsx"

            return Response(
                responsedata(True, "Template Downloaded", {"url": url}),
                status=status.HTTP_200_OK,
            )

        except:
            return Response(
                responsedata(False, "Can't save the template"),
                status=status.HTTP_400_BAD_REQUEST,
            )


class BulkStorage(APIView):
    """To create users from bulk upload template"""

    def post(self, request):

        # with open("bulkupload.xlsx","rb") as f:
        #     data = f.read()

        # data1 = base64.b64encode(data)
        # # f = open('file.txt', 'wb')
        # # f.write(data)
        # # f.close()

        # data2 = base64.b64decode(data1)
        try:
            data2 = base64.b64decode(request.data.get("data"))
        except:
            return Response(
                responsedata(False, "Something went wrong"),
                status=status.HTTP_400_BAD_REQUEST,
            )

        toread = io.BytesIO()
        toread.write(data2)
        toread.seek(0)
        df = pd.read_excel(toread)
        df.to_excel("output.xlsx")
        dfUsers = pd.read_excel("output.xlsx")

        failed_users = []
        successful_users = []

        # Processing each record
        for index, row in dfUsers.iterrows():
            firstName = row["First name"]
            lastName = row["Surname"]
            userName = row["Username"]
            if type(userName) == float:
                break
            email = row["Email"]
            customers = row["Customer"]
            role = row["Role"]

            # To check if company is missing
            try:
                customers = Company.objects.filter(name=customers).values().first()
                customers = customers.get("uuid")
            except:
                entry = {"Username": userName, "reason": "Missing Company Details"}
                failed_users.append(entry)
                continue

            # Autogenerated passwords
            password = randomString()

            # To check if any detail is missing
            flag = False
            data = {
                "firstName": firstName,
                "lastName": lastName,
                "userName": userName,
                "email": email,
                "role": role,
                "customers": customers,
                "password": password,
            }

            for value in data.values():
                try:
                    item = math.isnan(value)
                    if item:
                        flag = True
                        break
                except:
                    pass

            if flag:
                entry = {"Username": userName, "reason": "Missing Details"}
                failed_users.append(entry)
                continue

            if email_check(data["email"]):
                if not UserSettings.objects.filter(email=data["email"], customers=customers).exists():
                    ca_check = UserSettings.objects.filter(email=data["email"]).values().first()
                    if ca_check["role"] in ["projectadmin", "customeradmin"]:
                        entry = {"Username": userName, "reason": "Customer admin already exists"}
                        failed_users.append(entry)
                        continue
                    user = UserSettings.objects.get(email=data["email"])
                    user.customers.add(customers)
                    continue
                else:
                    entry = {"Username": userName, "reason": f"User already exists for {row['customer']}"}
                    failed_users.append(entry)
                    continue

            # Creating user in keycloak
            try:

                from main.fixer import find_keycloak_user, delete_keycloak_user
                exists = find_keycloak_user(email=data["email"])
                if exists is not False:
                    delete_keycloak_user(email=data["email"])
                    # Creating user in keycloak
                keyadmin = keycloak_admin.KeycloakAdmin(
                    server_url=deviare_settings.KEYCLOAK_URL,
                    username=deviare_settings.KEYCLOAK_ADMINUSER,
                    password=deviare_settings.KEYCLOAK_ADMINPASSWORD,
                    realm_name=deviare_settings.KEYCLOAK_REALM,
                )
                logger.info("Create KeyCloak User")
                logger.info(data)
                keyadmin.create_user(
                    {
                        "email": data["email"],
                        "enabled": True,
                        "username": data["userName"],
                        "credentials": [{"value": data["password"], "type": "password"}],
                        "realmRoles": ["user_default"],
                    }
                )
            except Exception as e:
                logger.exception(e)
            try:
                # Creating user in User Table
                logger.info("Create Auth User")
                user, c = User.objects.update_or_create(email=data["email"], username=data["userName"])
                user.set_password(data["password"])
                user.save()
                user.refresh_from_db()
                # Adding entry in UserSettings Table
                data["user"] = str(user.id)
            except Exception as e:
                logger.exception(e)
                entry = {"Username": userName, "reason": "User already exists"}
                failed_users.append(entry)
                pass

            # Adding entry in UserSettings Table
            try:
                logger.info("Create UserSettings")
                temp_password = data["password"]
                data["password"] = user.password
                data["customers"] = [data["customers"]]
                if data.get('user', None):
                    logger.info("Load User ID and PWD")
                    pk, pwd = User.objects.filter(email=data["email"]).values_list(
                        'pk', 'password').first()
                    data['user'] = pk
                    data["password"] = pwd
                user_qs = UserSettings.objects.filter(email=data["email"])
                if user_qs.exists():
                    user_obj = user_qs.first()
                    user_obj.set_password(data.pop('password'))
                    user_obj.save()
                    serializer = UserSettingsSerializer(user_obj, data=data)
                else:
                    serializer = UserSettingsSerializer(data=data)
                logger.info("Check Valid")
                if serializer.is_valid(raise_exception=True):
                    body1 = "Hi,\n\n\n\nYour account has been created, please find the login details below\n\n"
                    body2 = "Username : " + data["userName"] + "\n\n" + "Password : " + temp_password +  "\n\nURL: " + deviare_settings.URL +"/\n\n"
                    body3 = "Thanks, \n\n Deviare Support Team"
                    send_email(
                        [data["email"]],
                        "Deviare Login Details",
                        body1 + body2 + body3,
                    )
                    logger.info("Sent email")
                    serializer.save()
            except Exception as e:
                logger.exception(e)
                entry = {"Username": userName, "reason": "User already exists"}
                failed_users.append(entry)
                pass

            entry = {"Username": userName, "Role": data["role"], "Customer": row["Customer"]}
            successful_users.append(entry)

        result = {"Created": successful_users, "Failed": failed_users}
        len_created = len(result["Created"])
        len_failed = len(result["Failed"])
        return Response(
            responsedata(True, f"{len_created} Users created and {len_failed} failed", result),
            status=status.HTTP_200_OK,
        )


# Courses


class CourseList(generics.ListAPIView):
    """To list courses"""

    search_fields = ["name", "description", "provider"]
    filter_backends = (filters.SearchFilter,)

    def get_queryset(self):
        return Course.objects.all().order_by("name")

    def filter_queryset(self, queryset):
        for backend in list(self.filter_backends):
            queryset = backend().filter_queryset(self.request, queryset, self)
        return queryset

    def get(self, request):

        instance = self.filter_queryset(self.get_queryset())
        course_list = list(instance.values())

        pagenumber = request.GET.get("page", 1)
        paginator = Paginator(course_list, 10)

        if int(pagenumber) > paginator.num_pages:
            raise ValidationError("Not enough pages", code=404)
        try:
            previous_page_number = paginator.page(pagenumber).previous_page_number()
        except EmptyPage:
            previous_page_number = None
        try:
            next_page_number = paginator.page(pagenumber).next_page_number()
        except EmptyPage:
            next_page_number = None
        data_to_show = paginator.page(pagenumber).object_list

        return JsonResponse(
            {
                "pagination": {
                    "previous_page": previous_page_number,
                    "is_previous_page": paginator.page(pagenumber).has_previous(),
                    "next_page": next_page_number,
                    "is_next_page": paginator.page(pagenumber).has_next(),
                    "start_index": paginator.page(pagenumber).start_index(),
                    "end_index": paginator.page(pagenumber).end_index(),
                    "total_entries": paginator.count,
                    "total_pages": paginator.num_pages,
                    "page": int(pagenumber),
                },
                "results": data_to_show,
            },
            safe=False,
        )


def random_id():
    number = str(randint(1000000, 9999999))
    if not Course.objects.filter(course_id=number).exists():
        return number
    else:
        random_id()


class CourseCreate(APIView):
    """To create a course"""

    def post(self, request):
        if (
            request.data.get("name")
            and request.data.get("description")
            and request.data.get("provider")
            and request.data.get("link")
            and request.data.get("course_type")
        ):
            data = request.data
            if data["course_id"] == "":
                data["course_id"] = random_id()

            if Course.objects.filter(course_id=data["course_id"], course_type=data["course_type"]).exists():
                return Response(
                    responsedata(False, "Course exists with provided course id and course type"),
                    status=status.HTTP_400_BAD_REQUEST,
                )

            serializer = CourseSerializer(data=data)
            if serializer.is_valid(raise_exception=True):
                course = serializer.save()
                # Call task
                # create_course_on_talent_lms.apply_async((course.pk,),)
                return Response(
                    responsedata(True, f"{serializer.data['name']} course created successfully", serializer.data),
                    status=status.HTTP_200_OK,
                )
        else:
            return Response(responsedata(False, "Details not provided"), status=status.HTTP_400_BAD_REQUEST)


class CourseDetail(APIView):
    """Retrieve, update or delete a course"""

    def get(self, request, pk):
        course = Course.objects.filter(uuid=pk)
        serializer = CourseSerializer(course, many=True)
        return Response(
            responsedata(True, "Course retrieved successfully", serializer.data),
            status=status.HTTP_200_OK,
        )

    def put(self, request, pk):
        course = Course.objects.filter(uuid=pk).first()
        serializer = CourseSerializer(course, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(
                responsedata(True, "Course updated successfully", serializer.data),
                status=status.HTTP_200_OK,
            )
        return Response(
            responsedata(False, "Something went wrong", serializer.errors),
            status=status.HTTP_400_BAD_REQUEST,
        )

    def delete(self, request, pk):
        course = Course.objects.filter(uuid=pk)
        course.delete()
        return Response(responsedata(True, "Course Deleted"), status=status.HTTP_200_OK)


# Drop-Downs


# Countries List
def countrylist(request):
    return JsonResponse({"country_list": countries})


class CustomerDropDown(APIView):
    """To get list of customer names"""

    def get(self, request):

        if request.auth is None:
            return Response(responsedata(False, "You are Unauthorized"), status=status.HTTP_400_BAD_REQUEST)

        user = UserSettings.objects.filter(userName=request.user.username).values().first()

        if user["role"] == "superadmin":
            customers = list(Company.objects.filter().values("uuid", "name").order_by("name"))

        else:
            customers = list(
                UserSettings.objects.filter(userName=request.user.username).values_list("customers", flat=True)
            )
            customers = list(Company.objects.filter(uuid__in=customers).values("uuid", "name").order_by("name"))

        return Response(responsedata(True, "Customers List", customers), status=status.HTTP_200_OK)


class CustomerAdminDropDown(APIView):
    """To get list of customer admins as per selected customer"""

    def post(self, request):

        if request.data.get("customer"):

            customer_admin_list = list(
                UserSettings.objects.filter(customers=request.data.get("customer"),
                                            role__in=["projectadmin", "customeradmin"])
                .annotate(pa_projects=Count('allocated_project', distinct=True))
                .values("userName", "role", "uuid", "pa_projects")
                .order_by("userName")
            )

            if len(customer_admin_list) < 1:
                return Response(
                    responsedata(False, "Admins not available in the company"), status=status.HTTP_400_BAD_REQUEST
                )

            return Response(responsedata(True, "Customer Admins List", customer_admin_list), status=status.HTTP_200_OK)
        else:
            return Response(responsedata(False, "Customer Missing"), status=status.HTTP_400_BAD_REQUEST)


class CourseDropDown(APIView):
    """To get list of courses"""

    def get(self, request):

        courses = list(
            Course.objects.all().values("uuid", "name", "category", "course_id", "provider").order_by("name")
        )
        courses_list = []

        for item in courses:

            data = {
                "category": item["category"],
                "uuid": item["uuid"],
                "course": item["name"],
                "provider": item["provider"],
            }
            courses_list.append(data)

        return Response(responsedata(True, "Courses List", courses_list), status=status.HTTP_200_OK)


class AssessmentDropDown(APIView):
    """To get list of courses"""

    def get(self, request):

        assessments = list(
            Assessment.objects.all()
            .values(
                "uuid",
                "name",
                "description",
                "link",
            )
            .order_by("name")
        )
        return Response(responsedata(True, "Assessment List", assessments), status=status.HTTP_200_OK)


class ApprenticeshipDropDown(APIView):
    """To get list of Apprenticeship"""

    def get(self, request):

        apprenticeships = list(
            Apprenticeship.objects.all()
            .values(
                "uuid",
                "name",
                "description",
                "link",
            )
            .order_by("name")
        )
        return Response(responsedata(True, "Apprenticeship List", apprenticeships), status=status.HTTP_200_OK)

# Third-party APIs


class SimpliLearnCourses(APIView):
    """To save courses from simplilearn api in the database"""

    def get(self, request):

        # Simplilearn API for course catalog
        courses_url = "https://services.simplilearn.com/enterprise-catalog/get-course-catalog?country_code=US"
        headers = {"Authorization": "Bearer 5c166aee943ce89c741ccd8e6b8400e5"}
        all_courses = requests.get(courses_url, headers=headers)
        data = all_courses.json().get("data")
        count = 0

        # Saving into database
        for course in data:
            count += 1
            print(count)

            # Setting Blanks to Master Programs
            if course["category"] == "":
                course["category"] = "Master Programs"

            course_data = {
                "name": course["course_name"],
                "description": course["description"],
                "provider": "SimpliLearn",
                "link": course["course_url"],
                "category": course["category"],
                "course_id": course["course_id"],
                "course_type": course["course_type"],
            }

            # Update
            if Course.objects.filter(course_id=course["course_id"], course_type=course["course_type"]).exists():
                course = Course.objects.filter(
                    course_id=course["course_id"], course_type=course["course_type"]
                ).first()
                serializer = CourseSerializer(course, data=course_data, partial=True)
                if serializer.is_valid(raise_exception=True):
                    serializer.save()
            # Save
            else:
                serializer = CourseSerializer(data=course_data)
                if serializer.is_valid(raise_exception=True):
                    serializer.save()

        return Response(responsedata(True, "Courses Added to Database"), status=status.HTTP_200_OK)


class CourseData(APIView):
    def get(self, request):

        courses_df = pd.read_csv("course_data.csv", encoding="unicode_escape")

        count = 0
        for index, row in courses_df.iterrows():
            count += 1
            print(count)

            data = {}
            data["name"] = row["data__course_name"]
            data["description"] = row["data__description"]
            data["provider"] = "Simplilearn"
            data["link"] = row["data__course_url"]
            data["category"] = row["data__category"]
            data["course_id"] = row["data__course_id"]
            data["course_type"] = row["data__course_type"]

            serializer = CourseSerializer(data=data)
            if serializer.is_valid():
                serializer.save()

        return Response(responsedata(True, "Courses Added to Database"), status=status.HTTP_200_OK)


class MicrosoftCourses(APIView):
    def get(self, request):

        courses = requests.get("https://docs.microsoft.com/api/learn/catalog/")
        data = courses.json().get("learningPaths")
        count = 0
        continued = 0

        for item in data:
            course_data = {
                "name": item["title"],
                "description": item["summary"],
                "provider": "Microsoft",
                "link": item["url"],
                "category": "Microsoft",
                "course_id": random_id(),
                "course_type": "Microsoft",
            }

            if Course.objects.filter(name=item["title"], link=item["url"]).exists():
                continued += 1
                print(continued)
                print("continue")
                continue

            serializer = CourseSerializer(data=course_data)
            if serializer.is_valid():
                count += 1
                print(count)
                serializer.save()

        return Response(responsedata(True, "Courses Added to Database"), status=status.HTTP_200_OK)


class UpdateRecords(APIView):
    """Update reecords from Simplilearn data"""

    def post(self, request):

        failed_records = []
        no_changes = []
        count = 0

        df = pd.read_excel("userdata.xlsx", sheet_name="Data")

        for index, row in df.iterrows():

            if row["Learner Email"] == "" or row["Course Id"] == "" or row["Self-Leaning completion %"] == "":
                failed_records.append({"index": index, "reason": "Details missing"})
                continue

            if not UserSettings.objects.filter(email=row["Learner Email"]).exists():
                failed_records.append({"index": index, "reason": "User not available in database"})
                continue

            if not Course.objects.filter(course_id=row["Course Id"]).exists():
                failed_records.append({"index": index, "reason": "Course not available in database"})
                continue

            data = {
                "email": row["Learner Email"],
                "course_id": row["Course Id"],
                "course_completion": float(row["Self-Leaning completion %"]),
            }

            if not CourseLicenseUser.objects.filter(
                course_license_id__course_id__course_id=data["course_id"], user_id__email=data["email"]
            ).exists():
                failed_records.append({"index": index, "reason": "User not assigned to course on platform"})
                continue

            license = (
                CourseLicenseUser.objects.filter(
                    course_license_id__course_id__course_id=data["course_id"], user_id__email=data["email"]
                )
                .values("uuid", "course_completion")
                .first()
            )

            if license["course_completion"] != data["course_completion"]:

                license = CourseLicenseUser.objects.filter(uuid=license["uuid"]).first()
                serializer = CourseLicenseUserSerializer(
                    license, data={"course_completion": data["course_completion"]}, partial=True
                )

                if serializer.is_valid():
                    count += 1
                    print(count)
                    serializer.save()
                else:
                    failed_records.append({"index": index, "reason": serializer.errors})

            else:
                no_changes.append({"index": index, "Email": data["course_id"], "Course Id": data["course_id"]})
                continue

        data = {
            "updated_count": count,
            "no_changes_count": len(no_changes),
            "failed_records_count": len(failed_records),
            "no_changes": no_changes,
            "failed_records": failed_records,
        }

        return Response(responsedata(True, "Script finished successfully", data), status=status.HTTP_200_OK)


class CourseUserAssigmentAPIView(APIView):
    """
    Course user assignment operations
    """

    def post(self, request):
        """
        Add user to course
        """
        serializer = CourseUserAssignmentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = dict(serializer.validated_data)
        cu = CourseUserAssignment.objects.filter(course__pk=data["course_id"], user__email=data["email"])

        if cu.exists():
            # Assignment already exists
            return Response(responsedata(False, "Duplicate assignment"), status=status.HTTP_409_CONFLICT)
        else:
            user = get_object_or_404(UserSettings, email=data["email"])
            course = get_object_or_404(Course, pk=data["course_id"])

            CourseUserAssignment.objects.create(course=course, user=user)
            # Background task
            logger.info(data)
            assign_user_to_course_on_talent_lms.apply_async(kwargs=data)

            return Response(status=status.HTTP_200_OK)


class UserGoToCourseAPIView(APIView):
    """
    Course user retrieval
    """
    def api_call(self, endpoint='gotocourse', userinfo=None):
        from main.utils import add_user_to_course
        import pdb; pdb.set_trace()
        if userinfo is None:
            return False
        try:
            if endpoint == 'addusertocourse':
                return add_user_to_course(**userinfo)
            params = ','.join([f"{k}:{v}" for k, v in userinfo.items()])
            url = f"{URL_ADDRESS}/gotocourse/{params}"
            response = requests.request("GET", url, auth=(API_KEY, ""))
            if response.status_code == 200:
                return response.json()
            else:
                print(response.text)
        except Exception as exc:
            logger.exception(exc)
        return False

    def goto(self, userinfo):
        try:
            cl = userinfo.pop('course_license_id__project_id__company_id__branch')
            orig_user = userinfo.copy()
            userinfo.update(dict(
                logout_redirect=deviare_settings.URL,
                course_completed_redirect=deviare_settings.URL,
            ))

            content = self.api_call(endpoint='gotocourse', userinfo=userinfo)

            if content is False:
                added = self.api_call(endpoint='addusertocourse', userinfo=orig_user)
                if added is not False:
                    content = self.api_call(endpoint='gotocourse', userinfo=userinfo)
                else:
                    return Response(status=status.HTTP_404_NOT_FOUND)
            if content is not False:
                goto_url = content["goto_url"]
                if len(cl) > 0:
                    branch = f'{cl}.'
                    goto_url = content["goto_url"].replace('https://learning.deviare.co.za',
                                                       f"https://{branch}learning.deviare.co.za")
                return Response({"goto_url": goto_url}, status=status.HTTP_200_OK)
        except Exception as exc:
            logger.exception(exc)
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get(self, request, pk):
        """
        Return course link session url from TalentLMS
        """
        import pdb; pdb.set_trace()
        from main.utils import usersignup
        # course = get_object_or_404(Course, pk=pk)
        user = get_object_or_404(UserSettings, email=request.user.email)
        if not user.user_id_talentlms:
            user = usersignup(user)
        qs = CourseLicenseUser.objects.filter(
            course_license_id__course_id__pk=pk,
            user_id=user
        )
        if not qs.exists():
            return Response(status=status.HTTP_401_UNAUTHORIZED)
        try:
            userinfo = qs.annotate(
                taluser_id=F('user_id__user_id_talentlms'),
                talcourse_id=F('course_license_id__course_id__course_id_talent_lms'),
            ).values('taluser_id', 'talcourse_id', 'course_license_id__project_id__company_id__branch').first()
            userinfo['user_id'] = userinfo.pop('taluser_id')
            userinfo['course_id'] = userinfo.pop('talcourse_id')
            return self.goto(userinfo)
        except Exception as exc:
            logger.exception(exc)
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UserStatusInCourseAPIView(APIView):
    """
    Retrieve user progress status in course from
    Talent LMS
    """

    def post(self, request):
        serializer = CourseUserAssignmentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            API_KEY = "y1vJ4XRzdts9Kn1cAhMnKxRyy3VRCz"
            URL_ADDRESS = "https://deviare.talentlms.com/api/v1"

            data = dict(serializer.validated_data)

            user = UserSettings.objects.get(email=data["email"])
            course = Course.objects.get(pk=data["course_id"])

            user_id = user.user_id_talentlms
            course_id = course.course_id_talent_lms

            url = "%s/getuserstatusincourse/user_id:%s,course_id:%s" % (URL_ADDRESS, user_id, course_id)

            payload = {}
            headers = {}

            response = requests.request("GET", url, auth=(API_KEY, ""), headers=headers, data=payload)
            content = json.loads(response.text)
            logger.info(content)

            return Response({"course_progress_status": content}, status=status.HTTP_200_OK)

        except Exception as exc:
            logger.exception(exc)
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def user_progress(request):
    """
    Template view of user progress
    """

    return render(
        request,
        "reports/user_progress.html",
        {"course_name": "AWS Technical Essentials", "content": user_progress_data},
    )


def digital_readiness(request):
    """
    Template view of digital readiness
    """
    try:
        return render(request, "reports/digital_readiness.html", {"dimensions": digital_readiness_data})
    except Exception as exc:
        logger.critical(exc)
        return


class CustomerDigitalReadinessReport(APIView):

    @data_response
    def get(self, request, *args, **kwargs):
        """
        Returns digital readiness pdf report
        """
        d = {k: v[0] if type(v) in (list, tuple) else v for k, v in kwargs.items() if k in [ "user_id"]}

        serializer = CustomerDigitalReadinessReportSerializer(data=d)
        serializer.is_valid(raise_exception=True)
        data = serializer.data
        user = get_object_or_404(UserSettings, pk=data['user_id'])

        try:
            full_name = user.full_name
            reports = TMForumUserResponse.objects.filter(owner=user)
            dimensions = []
            file_path = os.path.join(deviare_settings.STATICFILES_DIRS[0], "images/%s.png" % (user.pk))
            if reports.exists():
                serializer = TMForumUserResponseDocSerializer(reports, many=True)
                data = serializer.data
                df = pd.DataFrame(data)
                df = df.groupby("dimension").mean()

                evaluation = df.to_dict()

                rating = {1: "Initiating", 2: "Emerging", 3: "Performing", 4: "Advancing", 5: "Leading"}

                for k, v in evaluation["aspiration_value"].items():
                    dimension = list(filter(lambda x: k in x['title'], digital_readiness_data))[0]
                    dimension["evaluation"] = rating[math.floor(v)]

                    dimensions.append(dimension)

                aspiration = [None] * 6
                status_quo = [None] * 6
                x_label = ["Customer", "Strategy", "Technology", "Operations", "Culture", "Data"]

                for k, v in evaluation["aspiration_value"].items():
                    index = x_label.index(k.capitalize())
                    aspiration.insert(index, math.floor(v))

                for k, v in evaluation["status_quo_value"].items():
                    index = x_label.index(k.capitalize())
                    status_quo.insert(index, math.floor(v))

                # Histogram on evaluation
                line_chart = pygal.Bar(style=BlueStyle)
                line_chart.title = "Digital Readiness Gap"
                line_chart.x_labels = ["Customer", "Strategy", "Technology", "Operations", "Culture", "Data"]
                line_chart.add("Aspiration", aspiration)
                line_chart.add("Status Quo", status_quo)

                line_chart.render_to_png(file_path)

            html_string = render_to_string(
                "reports/digital_readiness.html",
                {
                    "dimensions": dimensions,
                    "file_path": "images/%s.png" % (user.pk),
                    "full_name": full_name,
                    "created_at": datetime.datetime.now().strftime("%d %B %Y"),
                    "user_responses": False
                },
            )

            html = HTML(string=html_string, base_url=request.build_absolute_uri())
            pdf_file = html.write_pdf()

            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = f"attachment; filename=Digital_readiness_{full_name}.pdf"
            response["Content-Transfer-Encoding"] = "binary"

            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response.write(pdf_file)
            return response

        except Exception as exc:
            logger.exception(exc)
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UserLearningPathReport(APIView):

    @data_response
    def get(self, request,*args, **kwargs):
        """
        Return user learning path pdf report
        """
        from pprint import  pprint
        d = {k: v[0] if type(v) in (list, tuple) else v for k, v in kwargs.items() if k in ["project_id", "user_id"]}
        serializer = UserLearningPathReportSerializer(data=d)
        serializer.is_valid(raise_exception=True)
        data = serializer.data

        user = get_object_or_404(UserSettings, pk=data['user_id'])
        project = get_object_or_404(Project, pk=data["project_id"])
        try:
            # Retrieve user course licenses associated with project
            course_licenses = CourseLicenseUser.objects.filter(course_license_id__project_id=project.pk, user_id=user)

            learning_path = {}
            course_info = []
            learning_times = []
            for course_license in course_licenses:
                # Get report per user license
                entry = UserReport.objects.filter(user_license=course_license).first()
                if not entry:
                    continue

                # Need to classify according to category
                category = entry.user_license.course_license_id.course_id.category

                projects = []
                assessments = []
                live_classes = []
                # User report
                ur = entry.report

                k = list(ur.keys())
                d = ur

                if ur.get('Self-Learning Completion %') > 0.0:
                    course_info.append({'score': ur.get('Self-Learning Completion %'), 'course': ur.get('Course Name')})

                learning_times.append(ur.get('Self-Learning Time'))

                for i in range(0, len(k), 1):

                    if "Project Name" in k[i]:
                        if i + 1 < len(k):
                            if "Project Result" in k[i + 1]:
                                score = True if d[k[i + 1]].lower() == "passed" else False
                                projects.append({"title": d[k[i]], "status": score})

                    if "Assessment Test" in k[i]:
                        if i + 1 < len(k):
                            if "Best Score" in k[i + 1]:
                                assessments.append({"title": d[k[i]], "status": d[k[i + 1]]})

                    if "Exclusive Live Class" in k[i]:
                        if i + 2 < len(k):
                            if "ELVC" in k[i + 1]:
                                if "% of Sessions" in k[i + 2]:
                                    if d[k[i + 2]] > 0.0:
                                        live_classes.append({"title": k[i], "status": d[k[i + 2]]})
                content = []
                if projects:
                    content.append({"title": "Projects", "modules": projects, "average": 50, 'course_name': ur["Course Name"]})
                if assessments:
                    average = math.floor(sum(map(lambda x: x["status"], assessments)) / len(assessments))
                    # Filter out 0.0 values from assessments 
                    arr = list(filter(lambda x: x["status"] != 0.0, assessments))
                    if average > 0:
                        content.append(
                            {
                                "title": "Assessments",
                                "modules": arr,
                                "average": average,
                                'course_name': ur["Course Name"]
                            }
                        )
                if live_classes:
                    average = math.floor(sum(map(lambda x: x["status"], live_classes)) / len(live_classes))
                    # Filter out 0.0 values from live classes 
                    arr = list(filter(lambda x: x["status"] != 0.0, live_classes))
                    if average > 0:
                        content.append(
                            {
                                "title": "Exclusive Live Classes",
                                "modules": arr,
                                "average": average,
                                'course_name': ur["Course Name"]
                            }
                        )

                if not learning_path.get(category):
                    learning_path[category] = []

                # Append to category
                if content:
                    learning_path[category].append(content)


            learning_times = list(filter(lambda x: x != None, learning_times))
            total_learning_time = str(pd.Timedelta(seconds=sum(map(to_secs, learning_times))))
            course_info_average = 0
            if course_info:
                course_info_average = math.floor(sum(map(lambda x: x["score"], course_info)) / len(course_info))

            context = {
                'learning_path': learning_path,
                'full_name': user.full_name,
                'updated_at': datetime.datetime.now().strftime("%d %B %Y"),
                'course_info': course_info,
                'total_learning_time': total_learning_time,
                'course_info_average': course_info_average
            }

            html_string = render_to_string("reports/learning_path.html", context)
            html = HTML(string=html_string)
            pdf_file = html.write_pdf()

            response = HttpResponse(content_type="application/pdf;")
            response["Content-Disposition"] = "inline; filename=user_report.pdf"
            response["Content-Transfer-Encoding"] = "binary"
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response.write(pdf_file)
            return response


        except Exception as exc:
            logger.exception(exc)
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UserProgressReport(APIView):
    @data_response
    def get(self, request, *args, **kwargs):
        """
        Return user progress pdf report
        """
        d = {k: v[0] if type(v) in (list, tuple) else v for k, v in kwargs.items() if k in ["course_id", "user_id"]}

        serializer = UserProgressReportSerializer(data=d)
        serializer.is_valid(raise_exception=True)

        try:

            data = serializer.data
            course_license = CourseLicenseUser.objects.get(
                course_license_id__course_id__pk=data["course_id"], user_id=data["user_id"])
            entry = get_object_or_404(UserReport, user_license=course_license)

            ur = entry.report
            ul = entry.user_license

            full_name = ul.user_id.full_name

            k = list(ur.keys())
            d = ur

            projects = []
            assessments = []
            live_classes = []

            for i in range(0, len(k), 1):

                if "Project Name" in k[i]:
                    if i + 1 < len(k):
                        if "Project Result" in k[i + 1]:
                            score = True if d[k[i + 1]].lower() == "passed" else False
                            projects.append({"title": d[k[i]], "status": score})

                if "Assessment Test" in k[i]:
                    if i + 1 < len(k):
                        if "Best Score" in k[i + 1]:
                            assessments.append({"title": d[k[i]], "status": d[k[i + 1]]})
                        else:
                            assessments.append({"title": d[k[i]], "status": None})

                if "Exclusive Live Class" in k[i]:
                    if i + 2 < len(k):
                        if "ELVC" in k[i + 1]:
                            if "% of Sessions" in k[i + 2]:
                                live_classes.append({"title": k[i], "status": d[k[i + 2]]})
                        else:
                            live_classes.append({"title": k[i], "status": None})
            content = []
            if projects:
                content.append({"title": "Projects", "modules": projects, "average": 50})
            if assessments:
                content.append(
                    {
                        "title": "Assessments",
                        "modules": assessments,
                        "average": math.floor(sum(map(lambda x: x["status"], assessments)) / len(assessments)),
                    }
                )
            if live_classes:
                content.append({"title": "Exclusive Live Classes", "modules": live_classes, "average": 50})

            context = {
                "full_name": full_name,
                "course_name": ur["Course Name"],
                "learning_time": ur["Self-Learning Time"],
                "upated_at": entry.updated_at.strftime("%d %B %Y"),
                "content": content,
            }

            html_string = render_to_string("reports/user_progress.html", context)
            html = HTML(string=html_string)
            pdf_file = html.write_pdf()

            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = "inline; filename=user_report.pdf"
            response["Content-Transfer-Encoding"] = "binary"
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response.write(pdf_file)
            return response

        except CourseLicenseUser.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)


        except Exception as exc:
            logger.exception(exc)
            return dict(success=False, msg='No User Report for this user')


# Updating Records


def elearnStat_save_from_excel():

    # xl = pd.ExcelFile(excel_name)
    # sheet_name = xl.sheet_names

    # for i in range (len(sheet_name)):
    #     df = pd.read_excel(excel_name, sheet_name=sheet_name[i])
    #     if 'Self-Learning Completion %' in df.columns:
    #         break

    df = pd.read_csv("useractivity.csv")
    print("Total Records : ", len(df.index))

    failed_records = []
    no_changes = []
    count = 0

    for index, row in df.iterrows():

        if row["Learner Email"] == "" or row["Course Id"] == "" or row["Self-Learning Completion %"] == "":
            failed_records.append({"index": index, "reason": "Details missing"})
            print("Failed Record, Details missing. Total failed", len(failed_records))
            continue

        if not UserSettings.objects.filter(email=row["Learner Email"]).exists():
            failed_records.append({"index": index, "reason": "User not available in database"})
            print("Failed Record, User not available in database. Total failed", len(failed_records))
            continue

        if not Course.objects.filter(course_id=row["Course Id"]).exists():
            failed_records.append({"index": index, "reason": "Course not available in database"})
            print("Failed Record, Course not available in database. Total failed", len(failed_records))
            continue

        data = {
            "email": row["Learner Email"],
            "course_id": row["Course Id"],
            "course_completion": float(row["Self-Learning Completion %"]),
        }

        if not CourseLicenseUser.objects.filter(
            course_license_id__course_id__course_id=data["course_id"], user_id__email=data["email"]
        ).exists():
            failed_records.append({"index": index, "reason": "User not assigned to course on platform"})
            print("Failed Record, User not assigned to course on platform. Total failed", len(failed_records))
            continue

        license = (
            CourseLicenseUser.objects.filter(
                course_license_id__course_id__course_id=data["course_id"], user_id__email=data["email"]
            )
            .values("uuid", "course_completion")
            .first()
        )

        if license["course_completion"] != data["course_completion"]:

            license = CourseLicenseUser.objects.filter(uuid=license["uuid"]).first()
            serializer = CourseLicenseUserSerializer(
                license, data={"course_completion": data["course_completion"]}, partial=True
            )

            if serializer.is_valid():
                count += 1
                print("Data updated. Total", count)
                serializer.save()
            else:
                failed_records.append({"index": index, "reason": serializer.errors})
                print(f"Failed Record, {serializer.errors}. Total failed {len(failed_records)}")

        else:
            no_changes.append({"index": index, "Email": data["course_id"], "Course Id": data["course_id"]})
            print("No updates", len(no_changes))
            continue

    data = {
        "updated_count": count,
        "no_changes_count": len(no_changes),
        "failed_records_count": len(failed_records),
        "no_changes": no_changes,
        "failed_records": failed_records,
    }

    return data
    # return Response(responsedata(True, "Script finished successfully", data), status=status.HTTP_200_OK)


def preprocess_elearning_stat(messagesVal, access_token):

    for ml in messagesVal:

        if re.match(deviare_settings.deviare_config["SENDER_ADDRESS"], ml["from"]["emailAddress"]["address"]):
            data = ml["body"]["content"]
            soup = BeautifulSoup(data)

            for table in soup.findAll("table"):

                data = table.find("a")["href"]

                if "s3.amazonaws.com" in data and "Deviare_Learner_Activity" in data:

                    print("Downloading CSV")
                    csv_file = pd.read_csv(data)
                    print("Downloaded")
                    csv_file = csv_file.dropna(axis=1, how="all")
                    csv_file = csv_file.to_csv("useractivity.csv", index=False, header=True)

                    # S3 Upload
                    try:
                        s3 = boto3.client(
                            "s3",
                            aws_access_key_id=deviare_settings.deviare_config["AWS_S3_ACCESS_KEY"],
                            aws_secret_access_key=deviare_settings.deviare_config["AWS_S3_SECRET_KEY"],
                        )
                    except Exception as e:
                        print(str(e))
                        return Response(
                            responsedata(False, "Can't connect to Database"), status=status.HTTP_400_BAD_REQUEST
                        )

                    try:
                        with open("useractivity.csv", "rb") as f:
                            data = f.read()

                        data1 = base64.b64encode(data)
                        data2 = base64.b64decode(data1)
                        print("Uploading CSV")
                        s3.put_object(Body=data2, Bucket="elearn-stat", Key="tmp/useractivity.csv", ACL="public-read")
                        print("Uploaded")

                        # url = "https://elearn-stat.s3.amazonaws.com/tmp/useractivity.csv"

                    except Exception as e:
                        print(str(e))
                        return Response(responsedata(False, "Can't upload CSV"), status=status.HTTP_400_BAD_REQUEST)

                    # os.remove('useractivity.csv')

                    try:
                        data = elearnStat_save_from_excel()
                        return HttpResponse(json.dumps(data), content_type="application/json", status=200)
                    except Exception as e:
                        print(str(e))
                        return Response(
                            responsedata(False, "Something went wrong"), status=status.HTTP_400_BAD_REQUEST
                        )

    return Response(
        responsedata(False, "No matches found under current configuration"), status=status.HTTP_400_BAD_REQUEST
    )


def outlook_mail(request):
    logger.info("outlook mail")
    access_token = get_access_token(request, request.build_absolute_uri(reverse(read_outlook_mail)))
    messages = get_my_messages(access_token)
    return preprocess_elearning_stat(messages["value"], access_token)


def read_outlook_mail(request):
    logger.info("in read outlook mail")
    auth_code = request.GET["code"]
    redirect_uri = deviare_settings.deviare_config["MS_OAUTH_REDIRECT_URL"]
    token = get_token_from_code(auth_code, redirect_uri)

    access_token = token["access_token"]
    user = get_me(access_token)
    refresh_token = token["refresh_token"]
    expires_in = token["expires_in"]
    expiration = int(time.time()) + expires_in - 300

    request.session["access_token"] = access_token
    request.session["refresh_token"] = refresh_token
    request.session["token_expires"] = expiration

    return HttpResponseRedirect(reverse(outlook_mail))


def read_mail_api(request):
    try:
        logger.info(request)
        response_data = {}
        redirect_uri = deviare_settings.deviare_config["MS_OAUTH_REDIRECT_URL"]
        sign_in_url = get_signin_url(redirect_uri)
        return HttpResponseRedirect(sign_in_url)
    except Exception as e:
        response_data["ERROR"] = "Failed while oAuth for Microsoft Outlook - " + str(e)
        return HttpResponse(json.dumps(response_data), content_type="application/json", status=200)


def url_parse_plus(url):
    from urllib.parse import parse_qs
    burl, query = url.split('?')
    return dict(url=burl, params={k: v.pop() for k, v in parse_qs(query).items()})


class VGA(APIView):
    @data_response
    def get(self, request, *args, **kwargs):
        if not request.userAuth:
            return dict(status=False, msg='You are not logged in')
        p = request.session.get('raw_password', kwargs.get('session_data', []))
        if not p:
            return dict(status=False, msg='You are not current logged in')
        if type(p) in (list, tuple) and len(p) > 0:
            p = json.loads(base64.b64decode(p[0]).decode('utf-8')).get('password')

        from requests import Session
        import bs4
        default_launch = {
            "url": 'https://belong.auth.us-east-2.amazoncognito.com/oauth2/authorize',
            "params": {
                "identity_provider": 'DeviareOpenID',
                "redirect_uri": 'https://dashboard.belong.education/',
                'response_type': kwargs.get('response_type', 'TOKEN'),
                "client_id": '20738icdfn538kieuetqrajoj2',
                "scope": 'aws.cognito.signin.user.admin email openid profile',
            }
        }
        link = kwargs.get('link', False)
        try:
            if link is not False:
                if type(link) in (list, tuple) and len(link) > 0:
                    link = link[0]
                if type(link) == str:
                    if len(link) < 8:
                        link = default_launch
                    else:
                        link = url_parse_plus(link)
                else:
                    link = default_launch
            else:
                link = default_launch
        except Exception as e:
            logger.exception(e)
            link = default_launch
        link.update({"method": 'GET'})
        steps = {
            "launch": link,
            "login": {
                "method": "POST",
                "headers": {
                    "content-type": "application/x-www-form-urlencoded"
                },
                "data": {
                    "username": request.userAuth.userName,
                    "password": p
                }
            }
        }
        print(link)
        session = Session()
        response1 = session.request(**steps['launch'])
        if response1.url.startswith('https://identity.deviare'):
            htm = bs4.BeautifulSoup(response1.text)
            login_url = url_parse_plus(htm.find_all('form')[0].attrs['action'])
            if login_url['url'].startswith('/'):
                login_url['url'] = f"https://identity.deviare.co.za{login_url['url']}"
            steps['login'].update(login_url)
            response2 = session.request(**steps['login'])
            r = url_parse_plus(response2.history[0].url)
            return dict(url=f"{response2.url}&state={r.get('params').get('state')}")
        return dict(status=False, msg='We cannot complete your request. Please log out, log back in and try again')

def upload_csv():
        import pdb;
        import csv
        from .models import ElearninStates
        # pdb.set_trace()
        file = "Deviare_License_Learner_Activity_2021_09_14_EST.csv"


        filename = "Deviare_License_Learner_Activity_2021_09_14_EST.csv"
        dict_from_csv = pd.read_csv(filename, header=None, index_col=0, squeeze=True).to_dict()
        f = open(filename, 'r', encoding="utf-8")

        reader = csv.reader(f)
        lt = []
        for rw in reader:
            keys = rw
            break
        # Ignoring the header
        next(reader)
        # dict_from_csv = {rows[0]:rows[1] for rows in reader}
        for row in reader:
            print(f"row: {row}")
            new_dict = dict(zip(keys, row))
            user = ElearninStates.objects.get_or_create(email=keys['Learner Email'])
            user.name = keys['Learner Name']
            user.account_status = keys['Account Status']
            user.order_type = keys['Order Type']
            user.team = keys['Team']
            user.course_assignment_date = keys['Course Assignment Date']
            user.course_activation_date = keys['Course Activation Date']
            user.course_type = keys['Course Type']
            user.course_id = keys['Course Id']
            user.self_learning_completion = keys['Self-Learning Completion %']
            user.course_expiration_date = keys['Course Expiration Date']
            user.course_title = keys['Course Name']
            # user.test_score = keys['Course Type']
            # user.project_result = keys['Course Type']
            user.course_completion_date = keys['Certificate Unlock Date']
            user.live_class_attended = keys['Learning days']
            # user.osl_score = keys['Course Type']
            # user.lvc_sore = keys['Course Type']
            # user.project_score = keys['Course Type']
            user.assesment_test_score = keys['Course Type']
            # user.certification_score = keys['Course Type']
            user.concat = str(keys['Learner Email']) + str(keys['Course Id'])
            user.program = keys['Course Name']
            # user.certification_status = keys['Course Name']
            print(str(user.email) + " before")
            user.save()
            print(str(user.email) + " saved")


